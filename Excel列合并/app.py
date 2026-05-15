from flask import Flask, request, jsonify, send_file, render_template_string
import os
import uuid
import socket
import webbrowser
import threading
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
PROCESSED_FOLDER = os.path.join(BASE_DIR, 'processed')

for folder in [UPLOAD_FOLDER, PROCESSED_FOLDER]:
    os.makedirs(folder, exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

session_store = {}

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <title>Excel列合并工具</title>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }

        body {
            font-family: "Microsoft YaHei", Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }

        .container {
            max-width: 1100px;
            margin: 20px auto;
            background: rgba(255, 255, 255, 0.97);
            padding: 30px;
            border-radius: 16px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.15);
        }

        h1 {
            color: #333;
            text-align: center;
            margin-bottom: 8px;
            font-weight: 300;
            font-size: 2.2em;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .subtitle {
            text-align: center;
            color: #888;
            font-size: 0.95em;
            margin-bottom: 25px;
        }

        .info {
            background: #e3f2fd;
            padding: 15px 20px;
            border-radius: 10px;
            margin-bottom: 25px;
            border-left: 5px solid #2196F3;
        }

        .info ol { padding-left: 20px; }
        .info li { margin: 6px 0; line-height: 1.6; font-size: 0.93em; }

        .step-bar {
            display: flex;
            justify-content: center;
            gap: 0;
            margin-bottom: 30px;
        }

        .step-item {
            display: flex;
            align-items: center;
            font-size: 0.95em;
            color: #aaa;
            transition: all 0.3s;
        }

        .step-item.active {
            color: #667eea;
            font-weight: 600;
        }

        .step-item.done {
            color: #43e97b;
        }

        .step-num {
            width: 30px;
            height: 30px;
            border-radius: 50%;
            display: flex;
            align-items: center;
            justify-content: center;
            margin-right: 6px;
            font-weight: 700;
            font-size: 0.85em;
            border: 2px solid #ddd;
            color: #aaa;
            transition: all 0.3s;
        }

        .step-item.active .step-num {
            border-color: #667eea;
            background: #667eea;
            color: #fff;
        }

        .step-item.done .step-num {
            border-color: #43e97b;
            background: #43e97b;
            color: #fff;
        }

        .step-connector {
            width: 50px;
            height: 2px;
            background: #ddd;
            margin: 0 10px;
            align-self: center;
            transition: background 0.3s;
        }

        .step-connector.done { background: #43e97b; }

        .panel { display: none; }
        .panel.active { display: block; }

        .file-upload-wrapper {
            position: relative;
            overflow: hidden;
            border: 2px dashed #667eea;
            border-radius: 10px;
            background: #f8f9ff;
            transition: all 0.3s ease;
            padding: 30px 20px;
            text-align: center;
            cursor: pointer;
        }

        .file-upload-wrapper:hover,
        .file-upload-wrapper.dragover {
            border-color: #764ba2;
            background: #f0f4ff;
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.2);
        }

        .file-upload-wrapper i {
            font-size: 2.5em;
            color: #667eea;
            margin-bottom: 10px;
            display: block;
        }

        .file-upload-wrapper p { color: #666; font-size: 0.95em; }

        input[type="file"] {
            position: absolute;
            left: 0; top: 0;
            opacity: 0;
            width: 100%; height: 100%;
            cursor: pointer;
        }

        .file-list {
            max-height: 200px;
            overflow-y: auto;
            border: 1px solid #e0e0e0;
            border-radius: 10px;
            padding: 10px;
            background: #fafafa;
            margin-top: 12px;
        }

        .file-item {
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 8px 12px;
            background: white;
            border-radius: 6px;
            margin-bottom: 6px;
            box-shadow: 0 1px 3px rgba(0,0,0,0.08);
            font-size: 0.93em;
        }

        .file-item:last-child { margin-bottom: 0; }

        .file-info {
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .file-info i { color: #667eea; }

        .file-name {
            font-weight: 500;
            color: #333;
            max-width: 400px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }

        .file-size { font-size: 0.85em; color: #999; }

        .remove-file {
            background: none;
            border: none;
            color: #f5576c;
            cursor: pointer;
            font-size: 1em;
            padding: 4px 6px;
            border-radius: 4px;
            transition: all 0.2s;
        }

        .remove-file:hover { background: rgba(245, 87, 108, 0.1); }

        .btn {
            padding: 12px 28px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 50px;
            cursor: pointer;
            font-size: 15px;
            font-weight: 600;
            margin: 8px 6px;
            transition: all 0.3s ease;
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
            display: inline-flex;
            align-items: center;
            gap: 6px;
        }

        .btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
        }

        .btn:disabled {
            opacity: 0.5;
            cursor: not-allowed;
            transform: none;
            box-shadow: none;
        }

        .btn-success {
            background: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%);
            box-shadow: 0 4px 12px rgba(67, 233, 123, 0.4);
        }

        .btn-secondary {
            background: linear-gradient(135deg, #6c757d 0%, #495057 100%);
            box-shadow: 0 4px 12px rgba(108, 117, 125, 0.4);
        }

        .btn-danger {
            background: linear-gradient(135deg, #f5576c 0%, #f093fb 100%);
            box-shadow: 0 4px 12px rgba(245, 87, 108, 0.4);
        }

        .actions {
            text-align: center;
            margin-top: 25px;
        }

        .card {
            background: white;
            border: 1px solid #e8e8e8;
            border-radius: 10px;
            padding: 18px;
            margin-bottom: 16px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.04);
        }

        .card-header {
            display: flex;
            align-items: center;
            justify-content: space-between;
            margin-bottom: 12px;
        }

        .card-title {
            font-weight: 600;
            color: #333;
            font-size: 1em;
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .card-title i { color: #667eea; }

        .card-meta {
            font-size: 0.85em;
            color: #999;
        }

        .col-grid {
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
        }

        .col-chip {
            display: inline-flex;
            align-items: center;
            gap: 5px;
            padding: 6px 14px;
            border: 1px solid #ddd;
            border-radius: 20px;
            font-size: 0.88em;
            cursor: pointer;
            transition: all 0.2s;
            user-select: none;
            color: #555;
            background: #fafafa;
        }

        .col-chip:hover {
            border-color: #667eea;
            background: #f0f4ff;
        }

        .col-chip.selected {
            border-color: #667eea;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }

        .col-chip i {
            font-size: 0.8em;
        }

        .select-all-btn {
            font-size: 0.85em;
            color: #667eea;
            cursor: pointer;
            border: none;
            background: none;
            font-weight: 600;
            padding: 2px 8px;
        }

        .select-all-btn:hover { text-decoration: underline; }

        .preview-table-wrapper {
            overflow-x: auto;
            border: 1px solid #e8e8e8;
            border-radius: 8px;
            max-height: 400px;
            overflow-y: auto;
        }

        .preview-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.88em;
        }

        .preview-table th {
            background: #f8f9ff;
            padding: 10px 14px;
            text-align: left;
            font-weight: 600;
            color: #444;
            border-bottom: 2px solid #667eea;
            position: sticky;
            top: 0;
            z-index: 1;
            white-space: nowrap;
        }

        .preview-table td {
            padding: 8px 14px;
            border-bottom: 1px solid #eee;
            color: #555;
            max-width: 250px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }

        .preview-table tr:hover td { background: #f8f9ff; }

        .source-tag {
            display: inline-block;
            font-size: 0.75em;
            padding: 2px 8px;
            border-radius: 10px;
            background: #e3f2fd;
            color: #1565c0;
            margin-left: 6px;
        }

        .spinner {
            border: 3px solid rgba(102, 126, 234, 0.3);
            border-radius: 50%;
            border-top: 3px solid #667eea;
            width: 35px;
            height: 35px;
            animation: spin 1s linear infinite;
            margin: 0 auto 15px;
            display: none;
        }

        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }

        .loading-overlay {
            display: none;
            position: fixed;
            top: 0; left: 0; right: 0; bottom: 0;
            background: rgba(255,255,255,0.85);
            z-index: 999;
            justify-content: center;
            align-items: center;
            flex-direction: column;
        }

        .loading-overlay.active { display: flex; }

        .loading-overlay p {
            color: #667eea;
            font-weight: 600;
            font-size: 1.05em;
        }

        .result-box {
            display: none;
            padding: 20px;
            background: #e8f5e9;
            border-radius: 10px;
            border-left: 5px solid #4caf50;
            text-align: center;
        }

        .result-box h3 {
            color: #2e7d32;
            margin-bottom: 10px;
        }

        .result-box p { color: #555; margin-bottom: 15px; }

        .empty-hint {
            text-align: center;
            color: #bbb;
            padding: 30px;
            font-style: italic;
        }

        .merge-mode-selector {
            display: flex;
            gap: 12px;
            margin-bottom: 20px;
            justify-content: center;
        }

        .mode-option {
            padding: 12px 24px;
            border: 2px solid #ddd;
            border-radius: 10px;
            cursor: pointer;
            transition: all 0.3s;
            text-align: center;
            min-width: 180px;
        }

        .mode-option:hover { border-color: #667eea; }

        .mode-option.active {
            border-color: #667eea;
            background: #f0f4ff;
        }

        .mode-option i {
            display: block;
            font-size: 1.8em;
            color: #667eea;
            margin-bottom: 6px;
        }

        .mode-option .mode-title {
            font-weight: 600;
            color: #333;
            font-size: 0.95em;
        }

        .mode-option .mode-desc {
            font-size: 0.82em;
            color: #888;
            margin-top: 4px;
        }

        .sheet-selector {
            margin-bottom: 10px;
            font-size: 0.9em;
        }

        .sheet-selector label {
            margin-right: 8px;
            font-weight: 600;
            color: #555;
        }

        .sheet-selector select {
            padding: 4px 10px;
            border: 1px solid #ddd;
            border-radius: 6px;
            font-size: 0.9em;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1><i class="fas fa-table"></i> Excel列合并工具</h1>
        <p class="subtitle">上传多个Excel文件，选择需要的列，合并导出</p>

        <div class="info">
            <ol>
                <li>上传一个或多个 Excel 文件（.xlsx / .xls）</li>
                <li>系统自动识别每个文件的列标题，您勾选需要的列</li>
                <li>选择合并模式：纵向合并（数据追加）或横向合并（列拼接）</li>
                <li>预览合并结果并导出为新的 Excel 文件</li>
            </ol>
        </div>

        <div class="step-bar">
            <div class="step-item active" id="step1">
                <span class="step-num">1</span> 上传文件
            </div>
            <div class="step-connector" id="conn1"></div>
            <div class="step-item" id="step2">
                <span class="step-num">2</span> 选择列
            </div>
            <div class="step-connector" id="conn2"></div>
            <div class="step-item" id="step3">
                <span class="step-num">3</span> 预览导出
            </div>
        </div>

        <!-- Step 1: Upload -->
        <div class="panel active" id="panel1">
            <div class="file-upload-wrapper" id="uploadWrapper">
                <i class="fas fa-cloud-upload-alt"></i>
                <p>点击或拖拽 Excel 文件到此处上传（支持多选）</p>
                <input type="file" id="excelFiles" name="excel_files" multiple accept=".xlsx,.xls">
            </div>
            <div class="file-list" id="fileList">
                <div class="empty-hint">暂无文件，请上传 Excel 文件</div>
            </div>
            <div class="actions">
                <button class="btn btn-success" id="btnAnalyze" disabled onclick="analyzeFiles()">
                    <i class="fas fa-search"></i> 分析列信息
                </button>
                <button class="btn btn-secondary" onclick="clearFiles()">
                    <i class="fas fa-trash-alt"></i> 清空
                </button>
            </div>
        </div>

        <!-- Step 2: Select Columns -->
        <div class="panel" id="panel2">
            <div class="merge-mode-selector">
                <div class="mode-option active" id="modeVertical" onclick="setMergeMode('vertical')">
                    <i class="fas fa-arrows-alt-v"></i>
                    <div class="mode-title">纵向合并</div>
                    <div class="mode-desc">将选中列的数据逐行追加到一起</div>
                </div>
                <div class="mode-option" id="modeHorizontal" onclick="setMergeMode('horizontal')">
                    <i class="fas fa-arrows-alt-h"></i>
                    <div class="mode-title">横向合并</div>
                    <div class="mode-desc">将不同文件的列左右拼接到一起</div>
                </div>
            </div>

            <div id="columnCards"></div>

            <div class="actions">
                <button class="btn btn-secondary" onclick="goToStep(1)">
                    <i class="fas fa-arrow-left"></i> 上一步
                </button>
                <button class="btn btn-success" id="btnPreview" onclick="previewMerge()">
                    <i class="fas fa-eye"></i> 预览合并结果
                </button>
            </div>
        </div>

        <!-- Step 3: Preview & Export -->
        <div class="panel" id="panel3">
            <div class="preview-table-wrapper" id="previewWrapper">
                <table class="preview-table" id="previewTable">
                    <thead id="previewHead"></thead>
                    <tbody id="previewBody"></tbody>
                </table>
            </div>
            <div class="result-box" id="resultBox">
                <h3><i class="fas fa-check-circle"></i> 合并完成</h3>
                <p id="resultMsg"></p>
                <a href="" id="downloadLink" class="btn btn-success">
                    <i class="fas fa-download"></i> 下载 Excel 文件
                </a>
            </div>
            <div class="actions">
                <button class="btn btn-secondary" onclick="goToStep(2)">
                    <i class="fas fa-arrow-left"></i> 上一步
                </button>
                <button class="btn btn-success" id="btnExport" onclick="exportExcel()">
                    <i class="fas fa-file-export"></i> 导出 Excel
                </button>
            </div>
        </div>
    </div>

    <div class="loading-overlay" id="loadingOverlay">
        <div class="spinner" style="display:block;"></div>
        <p id="loadingText">处理中，请稍候...</p>
    </div>

    <script>
        let selectedFiles = [];
        let fileColumns = {};
        let mergeMode = 'vertical';
        let sessionId = '';
        let previewData = null;

        const uploadWrapper = document.getElementById('uploadWrapper');
        const fileInput = document.getElementById('excelFiles');

        uploadWrapper.addEventListener('dragover', function(e) {
            e.preventDefault();
            this.classList.add('dragover');
        });

        uploadWrapper.addEventListener('dragleave', function() {
            this.classList.remove('dragover');
        });

        uploadWrapper.addEventListener('drop', function(e) {
            e.preventDefault();
            this.classList.remove('dragover');
            if (e.dataTransfer.files.length > 0) {
                handleFiles(e.dataTransfer.files);
            }
        });

        fileInput.addEventListener('change', function() {
            if (this.files.length > 0) {
                handleFiles(this.files);
            }
        });

        function handleFiles(files) {
            for (let i = 0; i < files.length; i++) {
                const file = files[i];
                const ext = file.name.toLowerCase().split('.').pop();
                if (!['xlsx', 'xls'].includes(ext)) {
                    alert('文件 ' + file.name + ' 不是 Excel 格式，已跳过');
                    continue;
                }
                const isDup = selectedFiles.some(f => f.name === file.name && f.size === file.size);
                if (!isDup) {
                    selectedFiles.push(file);
                }
            }
            updateFileList();
        }

        function updateFileList() {
            const list = document.getElementById('fileList');
            if (selectedFiles.length === 0) {
                list.innerHTML = '<div class="empty-hint">暂无文件，请上传 Excel 文件</div>';
                document.getElementById('btnAnalyze').disabled = true;
                return;
            }
            document.getElementById('btnAnalyze').disabled = false;
            let html = '';
            selectedFiles.forEach((file, idx) => {
                const sizeMB = (file.size / (1024 * 1024)).toFixed(2);
                html += '<div class="file-item">' +
                    '<div class="file-info"><i class="fas fa-file-excel"></i>' +
                    '<span class="file-name">' + file.name + '</span>' +
                    '<span class="file-size">' + sizeMB + ' MB</span></div>' +
                    '<button class="remove-file" onclick="removeFile(' + idx + ')"><i class="fas fa-times"></i></button>' +
                    '</div>';
            });
            list.innerHTML = html;
        }

        function removeFile(idx) {
            selectedFiles.splice(idx, 1);
            updateFileList();
        }

        function clearFiles() {
            selectedFiles = [];
            fileColumns = {};
            updateFileList();
        }

        function showLoading(text) {
            document.getElementById('loadingText').textContent = text || '处理中，请稍候...';
            document.getElementById('loadingOverlay').classList.add('active');
        }

        function hideLoading() {
            document.getElementById('loadingOverlay').classList.remove('active');
        }

        function analyzeFiles() {
            if (selectedFiles.length === 0) {
                alert('请先上传 Excel 文件');
                return;
            }
            showLoading('正在分析列信息...');

            const formData = new FormData();
            selectedFiles.forEach(file => {
                formData.append('excel_files', file);
            });

            fetch('/analyze', {
                method: 'POST',
                body: formData
            })
            .then(res => res.json())
            .then(data => {
                hideLoading();
                if (data.status === 'success') {
                    sessionId = data.session_id;
                    fileColumns = data.files;
                    renderColumnCards();
                    goToStep(2);
                } else {
                    alert('分析失败: ' + data.message);
                }
            })
            .catch(err => {
                hideLoading();
                alert('请求失败: ' + err.message);
            });
        }

        function renderColumnCards() {
            const container = document.getElementById('columnCards');
            let html = '';
            const files = fileColumns;

            files.forEach((file, fIdx) => {
                html += '<div class="card">';
                html += '<div class="card-header">';
                html += '<div class="card-title"><i class="fas fa-file-excel"></i> ' + file.filename;
                html += '<span class="card-meta">（' + file.row_count + ' 行数据，' + file.columns.length + ' 列）</span></div>';
                html += '<div>';
                html += '<button class="select-all-btn" onclick="toggleAllCols(' + fIdx + ', true)">全选</button>';
                html += '<button class="select-all-btn" onclick="toggleAllCols(' + fIdx + ', false)">取消全选</button>';
                html += '</div></div>';

                if (file.sheets && file.sheets.length > 1) {
                    html += '<div class="sheet-selector">';
                    html += '<label>工作表：</label>';
                    html += '<select onchange="switchSheet(' + fIdx + ', this.value)">';
                    file.sheets.forEach((sh, sIdx) => {
                        html += '<option value="' + sIdx + '">' + sh + '</option>';
                    });
                    html += '</select></div>';
                }

                html += '<div class="col-grid" id="colGrid_' + fIdx + '">';
                file.columns.forEach((col, cIdx) => {
                    const selected = col.selected ? 'selected' : '';
                    html += '<div class="col-chip ' + selected + '" data-file="' + fIdx + '" data-col="' + cIdx + '" onclick="toggleCol(this, ' + fIdx + ', ' + cIdx + ')">';
                    html += '<i class="fas fa-check-circle"></i> ' + col.name;
                    html += '</div>';
                });
                html += '</div>';
                html += '</div>';
            });

            container.innerHTML = html;
        }

        function toggleCol(el, fIdx, cIdx) {
            fileColumns[fIdx].columns[cIdx].selected = !fileColumns[fIdx].columns[cIdx].selected;
            el.classList.toggle('selected');
        }

        function toggleAllCols(fIdx, selectAll) {
            fileColumns[fIdx].columns.forEach((col, cIdx) => {
                col.selected = selectAll;
            });
            const grid = document.getElementById('colGrid_' + fIdx);
            const chips = grid.querySelectorAll('.col-chip');
            chips.forEach(chip => {
                if (selectAll) chip.classList.add('selected');
                else chip.classList.remove('selected');
            });
        }

        function switchSheet(fIdx, sheetIdx) {
            showLoading('正在切换工作表...');
            fetch('/switch_sheet', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    session_id: sessionId,
                    file_index: fIdx,
                    sheet_index: parseInt(sheetIdx)
                })
            })
            .then(res => res.json())
            .then(data => {
                hideLoading();
                if (data.status === 'success') {
                    fileColumns[fIdx] = data.file_info;
                    renderColumnCards();
                } else {
                    alert('切换失败: ' + data.message);
                }
            })
            .catch(err => {
                hideLoading();
                alert('请求失败: ' + err.message);
            });
        }

        function setMergeMode(mode) {
            mergeMode = mode;
            document.getElementById('modeVertical').classList.toggle('active', mode === 'vertical');
            document.getElementById('modeHorizontal').classList.toggle('active', mode === 'horizontal');
        }

        function getSelection() {
            const selection = {};
            fileColumns.forEach((file, fIdx) => {
                const selectedCols = file.columns.filter(c => c.selected).map(c => c.name);
                if (selectedCols.length > 0) {
                    selection[fIdx] = selectedCols;
                }
            });
            return selection;
        }

        function previewMerge() {
            const selection = getSelection();
            if (Object.keys(selection).length === 0) {
                alert('请至少选择一个列');
                return;
            }
            showLoading('正在生成预览...');

            fetch('/preview', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    session_id: sessionId,
                    merge_mode: mergeMode,
                    selection: selection
                })
            })
            .then(res => res.json())
            .then(data => {
                hideLoading();
                if (data.status === 'success') {
                    previewData = data;
                    renderPreview(data);
                    goToStep(3);
                } else {
                    alert('预览失败: ' + data.message);
                }
            })
            .catch(err => {
                hideLoading();
                alert('请求失败: ' + err.message);
            });
        }

        function renderPreview(data) {
            const head = document.getElementById('previewHead');
            const body = document.getElementById('previewBody');

            let headerHtml = '<tr>';
            data.headers.forEach(h => {
                headerHtml += '<th>' + escapeHtml(h) + '</th>';
            });
            headerHtml += '</tr>';
            head.innerHTML = headerHtml;

            let bodyHtml = '';
            const maxRows = Math.min(data.rows.length, 100);
            for (let i = 0; i < maxRows; i++) {
                bodyHtml += '<tr>';
                data.rows[i].forEach(cell => {
                    bodyHtml += '<td title="' + escapeHtml(String(cell || '')) + '">' + escapeHtml(String(cell || '')) + '</td>';
                });
                bodyHtml += '</tr>';
            }
            if (data.rows.length > 100) {
                bodyHtml += '<tr><td colspan="' + data.headers.length + '" style="text-align:center;color:#999;font-style:italic;">仅显示前 100 行，共 ' + data.rows.length + ' 行数据</td></tr>';
            }
            body.innerHTML = bodyHtml;

            document.getElementById('resultBox').style.display = 'none';
        }

        function exportExcel() {
            const selection = getSelection();
            if (Object.keys(selection).length === 0) {
                alert('请至少选择一个列');
                return;
            }
            showLoading('正在导出...');

            fetch('/export', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify({
                    session_id: sessionId,
                    merge_mode: mergeMode,
                    selection: selection
                })
            })
            .then(res => res.json())
            .then(data => {
                hideLoading();
                if (data.status === 'success') {
                    document.getElementById('resultMsg').textContent =
                        '成功合并 ' + data.file_count + ' 个文件，共 ' + data.row_count + ' 行数据';
                    document.getElementById('downloadLink').href = data.download_url;
                    document.getElementById('resultBox').style.display = 'block';
                } else {
                    alert('导出失败: ' + data.message);
                }
            })
            .catch(err => {
                hideLoading();
                alert('请求失败: ' + err.message);
            });
        }

        function goToStep(step) {
            document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
            document.getElementById('panel' + step).classList.add('active');

            document.querySelectorAll('.step-item').forEach((s, i) => {
                s.classList.remove('active', 'done');
                if (i + 1 < step) s.classList.add('done');
                if (i + 1 === step) s.classList.add('active');
            });
            document.querySelectorAll('.step-connector').forEach((c, i) => {
                c.classList.toggle('done', i + 1 < step);
            });
        }

        function escapeHtml(str) {
            const div = document.createElement('div');
            div.textContent = str;
            return div.innerHTML;
        }
    </script>
</body>
</html>
'''


@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/analyze', methods=['POST'])
def analyze():
    try:
        files = request.files.getlist('excel_files')
        if not files:
            return jsonify({'status': 'error', 'message': '未上传文件'})

        sid = str(uuid.uuid4())
        session_store[sid] = {}

        upload_dir = os.path.join(UPLOAD_FOLDER, sid)
        os.makedirs(upload_dir, exist_ok=True)

        result = []
        for fIdx, f in enumerate(files):
            if not allowed_file(f.filename):
                continue
            filename = secure_filename(f.filename)
            if not filename:
                filename = f'file_{fIdx}.xlsx'
            filepath = os.path.join(upload_dir, filename)
            f.save(filepath)

            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheets = wb.sheetnames
            ws = wb.active

            columns = []
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row:
                for col_idx, val in enumerate(header_row, 1):
                    col_name = str(val).strip() if val else f'列{col_idx}'
                    columns.append({'name': col_name, 'index': col_idx, 'selected': True})

            row_count = 0
            for _ in ws.iter_rows(min_row=2, values_only=True):
                row_count += 1
            wb.close()

            file_info = {
                'filename': filename,
                'filepath': filepath,
                'sheets': sheets,
                'active_sheet': 0,
                'columns': columns,
                'row_count': row_count
            }
            result.append(file_info)
            session_store[sid][str(fIdx)] = file_info

        session_store[sid]['upload_dir'] = upload_dir

        return jsonify({'status': 'success', 'session_id': sid, 'files': result})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/switch_sheet', methods=['POST'])
def switch_sheet():
    try:
        data = request.json
        sid = data.get('session_id')
        fIdx = str(data.get('file_index'))
        sheet_index = data.get('sheet_index', 0)

        if sid not in session_store or fIdx not in session_store[sid]:
            return jsonify({'status': 'error', 'message': '会话不存在'})

        file_info = session_store[sid][fIdx]
        filepath = file_info['filepath']

        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames
        if sheet_index >= len(sheets):
            wb.close()
            return jsonify({'status': 'error', 'message': '工作表索引越界'})

        ws = wb.worksheets[sheet_index]
        columns = []
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            for col_idx, val in enumerate(header_row, 1):
                col_name = str(val).strip() if val else f'列{col_idx}'
                columns.append({'name': col_name, 'index': col_idx, 'selected': True})

        row_count = 0
        for _ in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
        wb.close()

        file_info['columns'] = columns
        file_info['row_count'] = row_count
        file_info['active_sheet'] = sheet_index

        return jsonify({'status': 'success', 'file_info': file_info})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/preview', methods=['POST'])
def preview():
    try:
        data = request.json
        sid = data.get('session_id')
        mode = data.get('merge_mode', 'vertical')
        selection = data.get('selection', {})

        if sid not in session_store:
            return jsonify({'status': 'error', 'message': '会话不存在'})

        headers, rows = do_merge(sid, mode, selection)
        session_store[sid]['_cached_merge'] = {'headers': headers, 'rows': rows}

        preview_rows = rows[:100]

        return jsonify({
            'status': 'success',
            'headers': headers,
            'rows': preview_rows,
            'total_rows': len(rows)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/export', methods=['POST'])
def export():
    try:
        data = request.json
        sid = data.get('session_id')
        mode = data.get('merge_mode', 'vertical')
        selection = data.get('selection', {})

        if sid not in session_store:
            return jsonify({'status': 'error', 'message': '会话不存在'})

        cached = session_store[sid].get('_cached_merge')
        if cached and cached.get('headers'):
            headers = cached['headers']
            rows = cached['rows']
        else:
            headers, rows = do_merge(sid, mode, selection)

        wb = Workbook()
        ws = wb.active
        ws.title = '合并结果'

        ws.append(headers)

        for row_data in rows:
            ws.append(row_data)

        col_widths = [0] * len(headers)
        for col_idx, h in enumerate(headers):
            col_widths[col_idx] = max(col_widths[col_idx], len(str(h)))
        for row_data in rows[:50]:
            for col_idx, val in enumerate(row_data):
                if val is not None:
                    col_widths[col_idx] = max(col_widths[col_idx], min(len(str(val)), 50))

        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width + 4

        out_dir = os.path.join(PROCESSED_FOLDER, sid)
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, '合并结果.xlsx')
        wb.save(out_path)

        file_count = len(selection)
        return jsonify({
            'status': 'success',
            'download_url': f'/download/{sid}',
            'file_count': file_count,
            'row_count': len(rows)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/download/<sid>')
def download(sid):
    out_path = os.path.join(PROCESSED_FOLDER, sid, '合并结果.xlsx')
    if not os.path.exists(out_path):
        return '文件不存在', 404
    return send_file(out_path, as_attachment=True, download_name='合并结果.xlsx')


def do_merge(sid, mode, selection):
    headers = []
    rows = []

    if mode == 'vertical':
        all_headers_set = set()
        ordered_headers = []

        for fIdx_str in sorted(selection.keys(), key=lambda x: int(x)):
            selected_col_names = selection[fIdx_str]
            for col_name in selected_col_names:
                if col_name not in all_headers_set:
                    all_headers_set.add(col_name)
                    ordered_headers.append(col_name)

        headers = ['来源文件'] + ordered_headers
        selected_col_names_set = set(ordered_headers)

        for fIdx_str in sorted(selection.keys(), key=lambda x: int(x)):
            file_info = session_store[sid][fIdx_str]
            filepath = file_info['filepath']
            sheet_index = file_info.get('active_sheet', 0)
            selected_col_names = set(selection[fIdx_str])

            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.worksheets[sheet_index]

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                wb.close()
                continue

            col_name_to_idx = {}
            for col_idx, val in enumerate(header_row, 1):
                col_name = str(val).strip() if val else f'列{col_idx}'
                col_name_to_idx[col_name] = col_idx - 1

            needed_indices = []
            for col_name in ordered_headers:
                if col_name in selected_col_names and col_name in col_name_to_idx:
                    needed_indices.append(col_name_to_idx[col_name])
                else:
                    needed_indices.append(None)

            for row_tuple in ws.iter_rows(min_row=2, values_only=True):
                row_data = [file_info['filename']]
                has_value = False
                for idx_pos, data_idx in enumerate(needed_indices):
                    if data_idx is not None and data_idx < len(row_tuple):
                        val = row_tuple[data_idx]
                        row_data.append(val)
                        if val is not None:
                            has_value = True
                    else:
                        row_data.append(None)
                if has_value:
                    rows.append(row_data)

            wb.close()

    elif mode == 'horizontal':
        for fIdx_str in sorted(selection.keys(), key=lambda x: int(x)):
            file_info = session_store[sid][fIdx_str]
            filename = file_info['filename']
            selected_col_names = selection[fIdx_str]

            for col_name in selected_col_names:
                headers.append(f'{filename}·{col_name}')

        all_file_rows = {}
        max_rows = 0

        for fIdx_str in sorted(selection.keys(), key=lambda x: int(x)):
            file_info = session_store[sid][fIdx_str]
            filepath = file_info['filepath']
            sheet_index = file_info.get('active_sheet', 0)
            selected_col_names = selection[fIdx_str]

            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.worksheets[sheet_index]

            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                wb.close()
                all_file_rows[fIdx_str] = []
                continue

            col_name_to_idx = {}
            for col_idx, val in enumerate(header_row, 1):
                col_name = str(val).strip() if val else f'列{col_idx}'
                col_name_to_idx[col_name] = col_idx - 1

            needed_indices = []
            for col_name in selected_col_names:
                needed_indices.append(col_name_to_idx.get(col_name))

            file_rows = []
            for row_tuple in ws.iter_rows(min_row=2, values_only=True):
                row_data = []
                for data_idx in needed_indices:
                    if data_idx is not None and data_idx < len(row_tuple):
                        row_data.append(row_tuple[data_idx])
                    else:
                        row_data.append(None)
                file_rows.append(row_data)

            wb.close()

            all_file_rows[fIdx_str] = file_rows
            max_rows = max(max_rows, len(file_rows))

        for i in range(max_rows):
            row_data = []
            for fIdx_str in sorted(selection.keys(), key=lambda x: int(x)):
                file_rows = all_file_rows.get(fIdx_str, [])
                if i < len(file_rows):
                    row_data.extend(file_rows[i])
                else:
                    row_data.extend([None] * len(selection[fIdx_str]))
            rows.append(row_data)

    return headers, rows


def open_browser(port):
    import time
    time.sleep(1.5)
    webbrowser.open(f'http://localhost:{port}')


if __name__ == '__main__':
    local_ip = get_local_ip()
    port = 5009

    print(f'\n{"="*50}')
    print(f'  Excel列合并工具 已启动')
    print(f'{"="*50}')
    print(f'  本机访问: http://localhost:{port}')
    print(f'  局域网访问: http://{local_ip}:{port}')
    print(f'{"="*50}\n')

    threading.Thread(target=open_browser, args=(port,), daemon=True).start()
    app.run(host='0.0.0.0', port=port, debug=False)
