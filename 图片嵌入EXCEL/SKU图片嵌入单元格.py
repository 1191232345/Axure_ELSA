from flask import Flask, request, jsonify, send_file, render_template_string
import os
import uuid
import shutil
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import socket

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-in-production'

# 设置上传文件大小限制为500MB
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024

# 获取当前工作目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
PROCESSED_FOLDER = os.path.join(BASE_DIR, 'processed')

print(f"项目根目录: {BASE_DIR}")
print(f"上传目录: {UPLOAD_FOLDER}")
print(f"处理目录: {PROCESSED_FOLDER}")

# 使用会话ID来隔离不同用户的数据
def get_user_upload_folder():
    session_id = str(uuid.uuid4())
    return os.path.join(UPLOAD_FOLDER, session_id)

def get_user_processed_folder():
    session_id = str(uuid.uuid4())
    return os.path.join(PROCESSED_FOLDER, session_id)

def ensure_directories():
    """确保所有必要的目录都存在"""
    try:
        if not os.path.exists(UPLOAD_FOLDER):
            os.makedirs(UPLOAD_FOLDER)
            print(f"✅ 创建上传目录: {UPLOAD_FOLDER}")
        else:
            print(f"✅ 上传目录已存在: {UPLOAD_FOLDER}")
        
        if not os.path.exists(PROCESSED_FOLDER):
            os.makedirs(PROCESSED_FOLDER)
            print(f"✅ 创建处理目录: {PROCESSED_FOLDER}")
        else:
            print(f"✅ 处理目录已存在: {PROCESSED_FOLDER}")
        
        return True
    except Exception as e:
        print(f"❌ 创建目录失败: {e}")
        return False

# 启动时确保目录存在
ensure_directories()

def get_local_ip():
    """获取本地IP地址"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

def is_image_file(filename):
    """检查文件是否为图片类型"""
    # 跳过Mac系统生成的隐藏文件（以._开头）
    if filename.startswith('._'):
        return False
    # 排除不支持的格式
    if filename.lower().endswith(('.mpo', '.mp')):
        return False
    return filename.lower().endswith(('.png', '.jpg', '.jpeg', '.jpge', '.gif', '.bmp'))

import re

def remove_chinese(text):
    """删除文本中的中文，只保留英文、数字和常见符号"""
    # 只保留 ASCII 字符（英文、数字、常见符号）
    return re.sub(r'[\u4e00-\u9fa5]', '', text)

def process_directories_and_generate_excel(directory_paths):
    """处理文件夹并生成Excel文件"""
    # 创建新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    
    # 设置表头
    ws.cell(row=1, column=1, value="SKU")
    
    # 确定最大图片数量，用于设置表头
    max_images = 0
    for dir_path in directory_paths:
        # 获取文件夹中的图片文件数量
        image_count = 0
        try:
            for file in os.listdir(dir_path):
                file_path = os.path.join(dir_path, file)
                if os.path.isfile(file_path) and is_image_file(file):
                    image_count += 1
            if image_count > max_images:
                max_images = image_count
        except Exception as e:
            print(f"统计图片数量错误 {dir_path}: {e}")
            continue
    
    # 设置图片列的表头
    for i in range(max_images):
        ws.cell(row=1, column=i+2, value=f"图片{i+1}")
    
    # 处理每个文件夹
    for row_idx, dir_path in enumerate(directory_paths, start=2):
        dir_name = os.path.basename(dir_path)
        # 删除中文，只保留英文、数字和常见符号
        dir_name = remove_chinese(dir_name)
        
        # 在第一列写入文件夹名称
        ws.cell(row=row_idx, column=1, value=dir_name)
        
        # 获取文件夹中的图片文件
        image_files = []
        for file in os.listdir(dir_path):
            file_path = os.path.join(dir_path, file)
            if os.path.isfile(file_path) and is_image_file(file):
                image_files.append(file_path)
        
        # 按顺序嵌入图片
        for col_idx, img_path in enumerate(image_files, start=2):
            try:
                # 调整列宽和行高
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
                ws.row_dimensions[row_idx].height = 70
                
                # 转换图片为PNG格式
                with PILImage.open(img_path) as pil_img:
                    # 创建内存中的BytesIO对象来存储PNG图片
                    png_buffer = BytesIO()
                    # 将图片转换为PNG格式并保存到缓冲区
                    pil_img.save(png_buffer, format='PNG')
                    png_buffer.seek(0)  # 重置缓冲区指针到开始位置
                    
                    # 使用转换后的PNG图片创建openpyxl Image对象
                    img = Image(png_buffer)
                    img.width = 80
                    img.height = 80
                    
                    # 将图片锚定到特定单元格
                    cell_address = f"{get_column_letter(col_idx)}{row_idx}"
                    img.anchor = cell_address
                    ws.add_image(img)
            except Exception as e:
                print(f"插入图片错误: {e}")
                continue
    
    # 保存Excel文件
    session_id = str(uuid.uuid4())
    output_dir = os.path.join(PROCESSED_FOLDER, session_id)
    os.makedirs(output_dir, exist_ok=True)
    output_excel_path = os.path.join(output_dir, f"图片汇总_{session_id}.xlsx")
    wb.save(output_excel_path)
    
    return output_excel_path

# HTML模板
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <title>图片嵌入Excel工具</title>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        :root {
            --primary-gradient: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            --secondary-gradient: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            --success-gradient: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%);
            --dark-gradient: linear-gradient(135deg, #2c3e50 0%, #3498db 100%);
            --danger-gradient: linear-gradient(135deg, #f5576c 0%, #f093fb 100%);
        }
        
        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
        }
        
        body { 
            font-family: "Microsoft YaHei", Arial, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        
        .container {
            max-width: 700px;
            margin: 20px auto;
            background: rgba(255, 255, 255, 0.95);
            padding: 25px;
            border-radius: 16px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.15);
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255,255,255,0.2);
        }
        
        h1 {
            color: #333;
            text-align: center;
            margin-bottom: 20px;
            font-weight: 300;
            font-size: 2.2em;
            background: var(--primary-gradient);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }
        
        .info {
            background: #e3f2fd;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 20px;
            border-left: 5px solid #2196F3;
        }
        
        .info ol {
            padding-left: 20px;
        }
        
        .info li {
            margin: 8px 0;
            line-height: 1.5;
            font-size: 0.95em;
        }
        
        .form-group { 
            margin: 20px 0;
        }
        
        label {
            display: block;
            margin-bottom: 10px;
            font-weight: 600;
            color: #444;
            font-size: 1em;
        }
        
        .file-upload-wrapper {
            position: relative;
            overflow: hidden;
            border: 2px dashed #667eea;
            border-radius: 10px;
            background: #f8f9ff;
            transition: all 0.3s ease;
            padding: 20px;
            text-align: center;
            cursor: pointer;
        }
        
        .file-upload-wrapper:hover {
            border-color: #764ba2;
            background: #f0f4ff;
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.2);
        }
        
        .file-upload-wrapper.dragover {
            border-color: #764ba2;
            background: #f0f4ff;
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.2);
        }
        
        .file-upload-wrapper i {
            font-size: 2.2em;
            color: #667eea;
            margin-bottom: 12px;
            display: block;
        }
        
        .file-upload-wrapper p {
            color: #666;
            margin-bottom: 12px;
            font-size: 0.95em;
        }
        
        input[type="file"] {
            position: absolute;
            left: 0;
            top: 0;
            opacity: 0;
            width: 100%;
            height: 100%;
            cursor: pointer;
        }
        
        .btn {
            padding: 12px 28px;
            background: var(--primary-gradient);
            color: white;
            border: none;
            border-radius: 50px;
            cursor: pointer;
            font-size: 16px;
            font-weight: 600;
            margin: 12px 8px;
            transition: all 0.3s ease;
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
            display: inline-flex;
            align-items: center;
            justify-content: center;
        }
        
        .btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
        }
        
        .btn:active {
            transform: translateY(-1px);
        }
        
        .btn-success {
            background: var(--success-gradient);
            box-shadow: 0 4px 12px rgba(67, 233, 123, 0.4);
        }
        
        .btn-danger {
            background: var(--danger-gradient);
            box-shadow: 0 4px 12px rgba(245, 87, 108, 0.4);
        }
        
        .btn i {
            margin-right: 6px;
            font-size: 0.9em;
        }
        
        .upload-result {
            margin-top: 20px;
            padding: 15px;
            background: #e8f5e8;
            border-radius: 10px;
            border-left: 5px solid #4caf50;
            display: none;
        }
        
        .upload-result h3 {
            margin-top: 0;
            color: #2e7d32;
            font-size: 1.2em;
        }
        
        .upload-result p {
            font-size: 0.95em;
            margin: 8px 0 0;
        }
        
        .spinner {
            border: 3px solid rgba(102, 126, 234, 0.3);
            border-radius: 50%;
            border-top: 3px solid #667eea;
            width: 35px;
            height: 35px;
            animation: spin 1s linear infinite;
            margin: 0 auto 15px;
            display: none;
        }
        
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        
        #progress {
            margin-top: 15px;
            display: none;
        }
        
        /* 文件列表样式 */
        .file-list {
            max-height: 180px;
            overflow-y: auto;
            border: 1px solid #e0e0e0;
            border-radius: 10px;
            padding: 12px;
            background: #fafafa;
        }
        
        .file-item {
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 8px 12px;
            background: white;
            border-radius: 6px;
            margin-bottom: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            transition: all 0.3s ease;
        }
        
        .file-item:hover {
            transform: translateX(4px);
            box-shadow: 0 3px 6px rgba(0,0,0,0.15);
        }
        
        .file-item:last-child {
            margin-bottom: 0;
        }
        
        .file-info {
            display: flex;
            align-items: center;
        }
        
        .file-info i {
            color: #667eea;
            margin-right: 8px;
            font-size: 1em;
        }
        
        .file-name {
            font-weight: 500;
            color: #333;
            max-width: 280px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            font-size: 0.95em;
        }
        
        .file-size {
            font-size: 0.85em;
            color: #666;
            margin-left: 8px;
        }
        
        .remove-file {
            background: none;
            border: none;
            color: #f5576c;
            cursor: pointer;
            font-size: 1em;
            padding: 4px;
            border-radius: 50%;
            transition: all 0.3s ease;
        }
        
        .remove-file:hover {
            background: rgba(245, 87, 108, 0.1);
            transform: scale(1.05);
        }
        
        /* 空状态提示 */
        .empty-state {
            text-align: center;
            color: #999;
            padding: 15px;
            font-style: italic;
            font-size: 0.95em;
        }
        
        /* 成功动画 */
        @keyframes fadeInUp {
            from {
                opacity: 0;
                transform: translateY(20px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        
        .upload-result {
            animation: fadeInUp 0.5s ease-out;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1><i class="fas fa-file-excel"></i> 图片嵌入Excel工具</h1>
        
        <div class="info">
            <ol>
                <li>支持上传常见压缩包格式（ZIP、RAR、7Z、TAR、GZ、TGZ、BZ2）</li>
                <li>自动解析文件夹</li>
                <li>文件夹内图片会自动排序，按文件名排序</li>
            </ol>
        </div>
        
        <form id="upload-form" enctype="multipart/form-data">
            <div class="form-group">
                <label for="zip_files"><i class="fas fa-file-archive"></i> 选择压缩包</label>
                <div class="file-upload-wrapper">
                    <i class="fas fa-cloud-upload-alt"></i>
                    <p>点击或拖拽压缩包到此处上传</p>
                    <input type="file" id="zip_files" name="zip_files" multiple accept=".zip,.rar,.7z,.tar,.gz,.tgz,.bz2">
                </div>
                
                <!-- 上传文件列表 -->
                <div id="file-list" class="file-list" style="margin-top: 15px;"></div>
            </div>
            
            <div style="text-align: center;">
                <button type="submit" class="btn btn-success">
                    <i class="fas fa-play"></i> 开始处理
                </button>
            </div>
        </form>
        
        <div id="progress">
            <div class="spinner"></div>
            <p style="text-align: center; color: #666;">正在处理，请稍候...</p>
        </div>
        
        <div class="upload-result" id="upload-result">
            <h3><i class="fas fa-check-circle"></i> 处理完成！</h3>
            <p id="result-message"></p>
            <div style="text-align: center; margin-top: 20px;">
                <a href="" id="download-link" class="btn">
                    <i class="fas fa-download"></i> 下载Excel文件
                </a>
            </div>
        </div>
    </div>
    
    <script>
        // 存储已选择的文件
        let selectedFiles = [];
        
        // 初始化
        document.addEventListener('DOMContentLoaded', function() {
            const fileInput = document.getElementById('zip_files');
            const fileList = document.getElementById('file-list');
            const form = document.getElementById('upload-form');
            const uploadWrapper = document.querySelector('.file-upload-wrapper');
            
            // 拖拽功能
            uploadWrapper.addEventListener('dragover', function(e) {
                e.preventDefault();
                this.classList.add('dragover');
            });
            
            uploadWrapper.addEventListener('dragleave', function() {
                this.classList.remove('dragover');
            });
            
            uploadWrapper.addEventListener('drop', function(e) {
                e.preventDefault();
                this.classList.remove('dragover');
                
                if (e.dataTransfer.files.length > 0) {
                    handleFiles(e.dataTransfer.files);
                }
            });
            
            // 文件选择事件
            fileInput.addEventListener('change', function() {
                if (this.files.length > 0) {
                    handleFiles(this.files);
                }
            });
            
            // 表单提交事件
            form.addEventListener('submit', function(e) {
                e.preventDefault();
                
                if (selectedFiles.length === 0) {
                    alert('请选择至少一个压缩包');
                    return;
                }
                
                // 创建FormData并添加所有选中的文件
                const formData = new FormData();
                selectedFiles.forEach(file => {
                    formData.append('zip_files', file);
                });
                
                // 显示进度
                document.getElementById('progress').style.display = 'block';
                document.getElementById('upload-result').style.display = 'none';
                
                // 发送请求
                fetch('/process', {
                    method: 'POST',
                    body: formData
                })
                .then(response => response.json())
                .then(data => {
                    document.getElementById('progress').style.display = 'none';
                    
                    if (data.status === 'success') {
                        document.getElementById('result-message').textContent = `成功处理了 ${data.directory_count} 个文件夹`;
                        document.getElementById('download-link').href = data.download_url;
                        document.getElementById('upload-result').style.display = 'block';
                    } else {
                        alert('处理失败: ' + data.message);
                    }
                })
                .catch(error => {
                    document.getElementById('progress').style.display = 'none';
                    alert('处理失败: ' + error.message);
                });
            });
        });
        
        // 处理文件选择
        function handleFiles(files) {
            for (let i = 0; i < files.length; i++) {
                const file = files[i];
                // 检查文件类型
                const ext = file.name.toLowerCase().substring(file.name.lastIndexOf('.'));
                if (!['.zip', '.rar', '.7z', '.tar', '.gz', '.tgz', '.bz2'].includes(ext)) {
                    alert(`文件 ${file.name} 不是支持的压缩包格式`);
                    continue;
                }
                // 检查是否已存在
                const isDuplicate = selectedFiles.some(f => f.name === file.name && f.size === file.size);
                if (!isDuplicate) {
                    selectedFiles.push(file);
                }
            }
            updateFileList();
        }
        
        // 更新文件列表显示
        function updateFileList() {
            const fileList = document.getElementById('file-list');
            
            if (selectedFiles.length === 0) {
                fileList.innerHTML = '<div class="empty-state">暂无上传的压缩包</div>';
                return;
            }
            
            let html = '';
            selectedFiles.forEach((file, index) => {
                html += `
                    <div class="file-item">
                        <div class="file-info">
                            <i class="fas fa-file-archive"></i>
                            <span class="file-name">${file.name}</span>
                            <span class="file-size">(${formatFileSize(file.size)})</span>
                        </div>
                        <button type="button" class="remove-file" onclick="removeFile(${index})">
                            <i class="fas fa-times"></i>
                        </button>
                    </div>
                `;
            });
            
            fileList.innerHTML = html;
        }
        
        // 删除文件
        function removeFile(index) {
            selectedFiles.splice(index, 1);
            updateFileList();
        }
        
        // 格式化文件大小
        function formatFileSize(bytes) {
            if (bytes < 1024) return bytes + ' B';
            else if (bytes < 1048576) return (bytes / 1024).toFixed(1) + ' KB';
            else if (bytes < 1073741824) return (bytes / 1048576).toFixed(1) + ' MB';
            else return (bytes / 1073741824).toFixed(1) + ' GB';
        }
    </script>
</body>
</html>
'''

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/process', methods=['POST'])
def process():
    try:
        # 获取上传的压缩包
        if 'zip_files' not in request.files:
            return jsonify({'status': 'error', 'message': '请选择压缩包'})
        
        uploaded_files = request.files.getlist('zip_files')
        
        if not uploaded_files:
            return jsonify({'status': 'error', 'message': '请选择至少一个压缩包'})
        
        # 创建临时上传目录
        temp_upload_dir = get_user_upload_folder()
        os.makedirs(temp_upload_dir, exist_ok=True)
        
        # 保存并解压所有压缩包
        all_directory_paths = []
        processed_skus = set()  # 用于记录已处理的SKU名称，防止重复处理
        
        for zip_file in uploaded_files:
            # 保存压缩包
            zip_file_path = os.path.join(temp_upload_dir, zip_file.filename)
            zip_file.save(zip_file_path)
            
            # 解压压缩包
            extract_dir = os.path.join(temp_upload_dir, os.path.splitext(zip_file.filename)[0])
            os.makedirs(extract_dir, exist_ok=True)
            
            # 解压逻辑
            if zip_file.filename.endswith('.zip'):
                import zipfile
                with zipfile.ZipFile(zip_file_path, 'r') as zf:
                    zf.extractall(extract_dir)
            elif zip_file.filename.endswith('.rar'):
                try:
                    import patoolib
                    patoolib.extract_archive(zip_file_path, outdir=extract_dir)
                except ImportError:
                    return jsonify({'status': 'error', 'message': '处理RAR文件需要安装patool库'})
            elif zip_file.filename.endswith('.7z'):
                try:
                    import py7zr
                    with py7zr.SevenZipFile(zip_file_path, mode='r') as zf:
                        zf.extractall(extract_dir)
                except ImportError:
                    return jsonify({'status': 'error', 'message': '处理7z文件需要安装py7zr库'})
            elif zip_file.filename.endswith('.tar'):
                import tarfile
                with tarfile.open(zip_file_path, 'r') as tf:
                    tf.extractall(extract_dir)
            elif zip_file.filename.endswith('.gz') or zip_file.filename.endswith('.tgz'):
                import tarfile
                with tarfile.open(zip_file_path, 'r:gz') as tf:
                    tf.extractall(extract_dir)
            elif zip_file.filename.endswith('.bz2'):
                import tarfile
                with tarfile.open(zip_file_path, 'r:bz2') as tf:
                    tf.extractall(extract_dir)
            else:
                return jsonify({'status': 'error', 'message': f'不支持的压缩格式: {os.path.splitext(zip_file.filename)[1]}'})
            
            # 遍历解压后的目录，收集所有包含图片的文件夹
            for root, dirs, files in os.walk(extract_dir):
                # 跳过Mac系统生成的__MACOSX文件夹
                if '__MACOSX' in root:
                    continue
                try:
                    # 检查目录中是否有图片文件
                    has_images = any(is_image_file(file) for file in files)
                    sku_name = os.path.basename(root)  # 使用文件夹名称作为SKU
                    # 删除中文，只保留英文、数字和常见符号
                    sku_name = remove_chinese(sku_name)
                    if has_images and sku_name not in processed_skus:
                        all_directory_paths.append(root)
                        processed_skus.add(sku_name)
                except Exception as e:
                    print(f"处理目录错误 {root}: {e}")
                    continue
        
        # 处理文件夹并生成Excel
        output_excel_path = process_directories_and_generate_excel(all_directory_paths)
        
        # 生成下载链接
        download_url = f"/download?file={os.path.basename(output_excel_path)}"
        
        return jsonify({
            'status': 'success',
            'directory_count': len(all_directory_paths),
            'download_url': download_url
        })
        
    except Exception as e:
        print(f"处理错误: {e}")
        return jsonify({'status': 'error', 'message': f'处理失败: {str(e)}'})

@app.route('/download')
def download():
    """下载生成的Excel文件"""
    try:
        file_name = request.args.get('file')
        
        if not file_name:
            return jsonify({'status': 'error', 'message': '文件名不能为空'})
        
        # 查找文件
        file_path = None
        for root, dirs, files in os.walk(PROCESSED_FOLDER):
            if file_name in files:
                file_path = os.path.join(root, file_name)
                break
        
        if not file_path:
            return jsonify({'status': 'error', 'message': '文件不存在'})
        
        # 发送文件
        return send_file(file_path, as_attachment=True, download_name=file_name)
        
    except Exception as e:
        print(f"下载错误: {e}")
        return jsonify({'status': 'error', 'message': f'下载失败: {str(e)}'})

if __name__ == '__main__':
    local_ip = get_local_ip()
    print(f"\n✅ 服务器启动成功！")
    print(f"✅ 访问地址: http://{local_ip}:5002")
    print(f"✅ 或访问: http://127.0.0.1:5002\n")
    app.run(host='0.0.0.0', port=5002, debug=True)