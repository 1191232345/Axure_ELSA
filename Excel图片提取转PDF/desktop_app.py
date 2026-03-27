import sys
import os
import threading
import webbrowser
import socket
import tkinter as tk
from tkinter import ttk, messagebox
import time
import traceback

from flask import Flask, request, jsonify, send_file, render_template_string
import uuid
import shutil
from io import BytesIO
from openpyxl import load_workbook
from PIL import Image as PILImage
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import zipfile

def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

def find_available_port(start_port=5004, max_attempts=10):
    for port in range(start_port, start_port + max_attempts):
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.bind(('', port))
                return port
        except OSError:
            continue
    return start_port

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
PROCESSED_FOLDER = os.path.join(BASE_DIR, 'processed')

def ensure_directories():
    try:
        if not os.path.exists(UPLOAD_FOLDER):
            os.makedirs(UPLOAD_FOLDER)
        if not os.path.exists(PROCESSED_FOLDER):
            os.makedirs(PROCESSED_FOLDER)
        return True
    except Exception as e:
        print(f"创建目录失败: {e}")
        return False

ensure_directories()

def extract_tracking_number(excel_path):
    try:
        wb = load_workbook(excel_path, data_only=True)
        if len(wb.worksheets) == 0:
            return None
        ws = wb.worksheets[0]
        for row_idx in range(1, min(100, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and 'tracking' in str(cell_value).lower():
                    tracking_value = ws.cell(row=row_idx, column=col_idx + 1).value
                    if tracking_value:
                        wb.close()
                        return str(tracking_value).strip()
        wb.close()
        return None
    except Exception as e:
        print(f"提取Tracking Number失败: {e}")
        return None

def extract_and_compress_images(excel_path, sheet_indices=[2, 3, 4, 5], max_image_size=2000):
    images_data = []
    try:
        wb = load_workbook(excel_path)
        for sheet_idx in sheet_indices:
            if sheet_idx >= len(wb.worksheets):
                continue
            ws = wb.worksheets[sheet_idx]
            sheet_name = ws.title
            if hasattr(ws, '_images'):
                img_count = len(ws._images)
                print(f"正在处理Sheet '{sheet_name}': 提取{img_count}张图片...")
                for img_idx, img in enumerate(ws._images):
                    try:
                        print(f"  提取图片 {img_idx + 1}/{img_count}...", end='', flush=True)
                        image_data = img._data()
                        pil_image = PILImage.open(BytesIO(image_data))
                        original_size = pil_image.size
                        if pil_image.mode in ('RGBA', 'LA', 'P'):
                            background = PILImage.new('RGB', pil_image.size, (255, 255, 255))
                            if pil_image.mode == 'P':
                                pil_image = pil_image.convert('RGBA')
                            background.paste(pil_image, mask=pil_image.split()[-1] if pil_image.mode == 'RGBA' else None)
                            pil_image = background
                        if max(pil_image.size) > max_image_size:
                            ratio = max_image_size / max(pil_image.size)
                            new_size = (int(pil_image.size[0] * ratio), int(pil_image.size[1] * ratio))
                            pil_image = pil_image.resize(new_size, PILImage.Resampling.LANCZOS)
                        compressed_buffer = BytesIO()
                        pil_image.save(compressed_buffer, format='JPEG', quality=85, optimize=True)
                        compressed_buffer.seek(0)
                        img_width = img.width
                        img_height = img.height
                        anchor = img.anchor
                        if hasattr(anchor, '_from'):
                            cell_ref = f"{anchor._from.col}{anchor._from.row + 1}"
                        else:
                            cell_ref = "Unknown"
                        images_data.append({
                            'sheet_name': sheet_name,
                            'sheet_index': sheet_idx + 1,
                            'image_data': compressed_buffer.getvalue(),
                            'width': img_width,
                            'height': img_height,
                            'cell': cell_ref,
                            'original_size': original_size,
                            'compressed_size': pil_image.size
                        })
                        print(f" 完成 ({original_size[0]}x{original_size[1]} -> {pil_image.size[0]}x{pil_image.size[1]})")
                    except Exception as e:
                        print(f" 失败")
                        print(f"  提取图片失败: {e}")
                        continue
        wb.close()
    except Exception as e:
        print(f"处理Excel文件失败: {e}")
    return images_data

def create_pdf_from_images(images_data, output_path, excel_filename, max_pdf_size_mb=4):
    if not images_data:
        return False
    c = canvas.Canvas(output_path, pagesize=A4)
    page_width, page_height = A4
    y_position = page_height - 50
    current_sheet = None
    for idx, img_data in enumerate(images_data):
        if img_data['sheet_name'] != current_sheet:
            if current_sheet is not None:
                c.showPage()
                y_position = page_height - 50
            else:
                y_position = page_height - 50
            current_sheet = img_data['sheet_name']
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, y_position, f"{current_sheet}")
            y_position -= 30
        img_width = img_data['width']
        img_height = img_data['height']
        max_width = page_width - 100
        max_height = 500
        if img_width > max_width or img_height > max_height:
            ratio = min(max_width / img_width, max_height / img_height)
            img_width = img_width * ratio
            img_height = img_height * ratio
        if y_position - img_height < 50:
            c.showPage()
            y_position = page_height - 50
        image_data = img_data['image_data']
        if image_data:
            temp_img_path = os.path.join(UPLOAD_FOLDER, f"temp_img_{uuid.uuid4()}.jpg")
            with open(temp_img_path, 'wb') as f:
                f.write(image_data)
            c.drawImage(temp_img_path, 50, y_position - img_height, width=img_width, height=img_height)
            try:
                os.remove(temp_img_path)
            except:
                pass
        c.setFont("Helvetica", 10)
        c.drawString(50, y_position - img_height - 15, f"Cell: {img_data['cell']}")
        y_position -= (img_height + 40)
    c.save()
    pdf_size = os.path.getsize(output_path)
    if pdf_size > max_pdf_size_mb * 1024 * 1024:
        print(f"警告: PDF文件大小 {pdf_size / 1024 / 1024:.2f}MB 超过限制 {max_pdf_size_mb}MB")
    else:
        print(f"PDF文件大小: {pdf_size / 1024 / 1024:.2f}MB")
    return True

def process_excel_files(excel_files):
    session_id = str(uuid.uuid4())
    output_dir = os.path.join(PROCESSED_FOLDER, session_id)
    os.makedirs(output_dir, exist_ok=True)
    all_results = []
    for excel_file in excel_files:
        filename = os.path.basename(excel_file)
        tracking_number = extract_tracking_number(excel_file)
        images_data = extract_and_compress_images(excel_file)
        if images_data:
            if tracking_number:
                pdf_filename = f"{tracking_number}.pdf"
            else:
                pdf_filename = f"{os.path.splitext(filename)[0]}_images.pdf"
            pdf_path = os.path.join(output_dir, pdf_filename)
            success = create_pdf_from_images(images_data, pdf_path, filename)
            if success:
                all_results.append({
                    'filename': filename,
                    'pdf_filename': pdf_filename,
                    'image_count': len(images_data),
                    'tracking_number': tracking_number,
                    'status': 'success'
                })
            else:
                all_results.append({
                    'filename': filename,
                    'status': 'failed',
                    'message': 'PDF生成失败'
                })
        else:
            all_results.append({
                'filename': filename,
                'status': 'no_images',
                'message': '未找到图片'
            })
    return all_results, session_id

HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <title>Excel图片提取转PDF工具</title>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.0.0/css/all.min.css" rel="stylesheet">
    <style>
        :root {
            --primary-gradient: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            --secondary-gradient: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
            --success-gradient: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%);
            --danger-gradient: linear-gradient(135deg, #f5576c 0%, #f093fb 100%);
        }
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { 
            font-family: "Microsoft YaHei", Arial, sans-serif; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }
        .container {
            max-width: 800px;
            margin: 20px auto;
            background: rgba(255, 255, 255, 0.95);
            padding: 30px;
            border-radius: 16px;
            box-shadow: 0 20px 40px rgba(0,0,0,0.15);
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255,255,255,0.2);
        }
        h1 {
            color: #333;
            text-align: center;
            margin-bottom: 25px;
            font-weight: 300;
            font-size: 2.2em;
            background: var(--primary-gradient);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }
        .info {
            background: #e3f2fd;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 25px;
            border-left: 5px solid #2196F3;
        }
        .info ol { padding-left: 20px; }
        .info li {
            margin: 8px 0;
            line-height: 1.6;
            font-size: 0.95em;
        }
        .form-group { margin: 25px 0; }
        label {
            display: block;
            margin-bottom: 12px;
            font-weight: 600;
            color: #444;
            font-size: 1em;
        }
        .file-upload-wrapper {
            position: relative;
            overflow: hidden;
            border: 2px dashed #667eea;
            border-radius: 12px;
            background: #f8f9ff;
            transition: all 0.3s ease;
            padding: 40px 20px;
            text-align: center;
            cursor: pointer;
        }
        .file-upload-wrapper:hover {
            border-color: #764ba2;
            background: #f0f4ff;
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.2);
        }
        .file-upload-wrapper.dragover {
            border-color: #764ba2;
            background: #f0f4ff;
            box-shadow: 0 8px 20px rgba(102, 126, 234, 0.2);
        }
        .file-upload-wrapper i {
            font-size: 3em;
            color: #667eea;
            margin-bottom: 15px;
            display: block;
        }
        .file-upload-wrapper p {
            color: #666;
            margin-bottom: 10px;
            font-size: 1em;
        }
        input[type="file"] {
            position: absolute;
            left: 0;
            top: 0;
            opacity: 0;
            width: 100%;
            height: 100%;
            cursor: pointer;
        }
        .btn {
            padding: 14px 32px;
            background: var(--primary-gradient);
            color: white;
            border: none;
            border-radius: 50px;
            cursor: pointer;
            font-size: 16px;
            font-weight: 600;
            margin: 15px 10px;
            transition: all 0.3s ease;
            box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4);
            display: inline-flex;
            align-items: center;
            justify-content: center;
        }
        .btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
        }
        .btn:active { transform: translateY(-1px); }
        .btn-success {
            background: var(--success-gradient);
            box-shadow: 0 4px 12px rgba(67, 233, 123, 0.4);
        }
        .btn-danger {
            background: var(--danger-gradient);
            box-shadow: 0 4px 12px rgba(245, 87, 108, 0.4);
        }
        .btn i {
            margin-right: 8px;
            font-size: 1em;
        }
        .file-list {
            max-height: 250px;
            overflow-y: auto;
            border: 1px solid #e0e0e0;
            border-radius: 10px;
            padding: 12px;
            background: #fafafa;
            margin-top: 15px;
        }
        .file-item {
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 10px 14px;
            background: white;
            border-radius: 8px;
            margin-bottom: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            transition: all 0.3s ease;
        }
        .file-item:hover {
            transform: translateX(4px);
            box-shadow: 0 3px 6px rgba(0,0,0,0.15);
        }
        .file-item:last-child { margin-bottom: 0; }
        .file-info {
            display: flex;
            align-items: center;
            flex: 1;
        }
        .file-info i {
            color: #667eea;
            margin-right: 10px;
            font-size: 1.2em;
        }
        .file-name {
            font-weight: 500;
            color: #333;
            max-width: 400px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            font-size: 0.95em;
        }
        .file-size {
            font-size: 0.85em;
            color: #666;
            margin-left: 10px;
        }
        .remove-file {
            background: none;
            border: none;
            color: #f5576c;
            cursor: pointer;
            font-size: 1.2em;
            padding: 6px;
            border-radius: 50%;
            transition: all 0.3s ease;
        }
        .remove-file:hover {
            background: rgba(245, 87, 108, 0.1);
            transform: scale(1.1);
        }
        .empty-state {
            text-align: center;
            color: #999;
            padding: 20px;
            font-style: italic;
            font-size: 0.95em;
        }
        .progress {
            margin-top: 20px;
            display: none;
        }
        .spinner {
            border: 3px solid rgba(102, 126, 234, 0.3);
            border-radius: 50%;
            border-top: 3px solid #667eea;
            width: 40px;
            height: 40px;
            animation: spin 1s linear infinite;
            margin: 0 auto 15px;
        }
        @keyframes spin {
            0% { transform: rotate(0deg); }
            100% { transform: rotate(360deg); }
        }
        .result {
            margin-top: 25px;
            display: none;
        }
        .result-card {
            background: #e8f5e8;
            padding: 20px;
            border-radius: 10px;
            border-left: 5px solid #4caf50;
            margin-bottom: 15px;
            animation: fadeInUp 0.5s ease-out;
        }
        .result-card.error {
            background: #ffebee;
            border-left-color: #f44336;
        }
        .result-card.warning {
            background: #fff3e0;
            border-left-color: #ff9800;
        }
        .result-title {
            font-size: 1.1em;
            font-weight: 600;
            color: #2e7d32;
            margin-bottom: 10px;
        }
        .result-card.error .result-title { color: #c62828; }
        .result-card.warning .result-title { color: #e65100; }
        .result-item {
            padding: 8px 0;
            border-bottom: 1px solid rgba(0,0,0,0.1);
        }
        .result-item:last-child { border-bottom: none; }
        .download-section {
            text-align: center;
            margin-top: 20px;
        }
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(20px); }
            to { opacity: 1; transform: translateY(0); }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1><i class="fas fa-file-pdf"></i> Excel图片提取转PDF工具</h1>
        <div class="info">
            <ol>
                <li>支持拖拽上传一个或多个Excel文件(.xlsx, .xls)</li>
                <li>自动提取Invoice、Customer Message、Damage Photo和SupportingDocuments四个Sheet页中的图片</li>
                <li>将提取的图片嵌入PDF文件并导出</li>
                <li>每个Excel文件生成一个独立的PDF文件</li>
                <li>自动压缩图片，确保PDF文件不超过4MB</li>
            </ol>
        </div>
        <form id="upload-form" enctype="multipart/form-data">
            <div class="form-group">
                <label for="excel_files"><i class="fas fa-file-excel"></i> 选择Excel文件</label>
                <div class="file-upload-wrapper" id="drop-zone">
                    <i class="fas fa-cloud-upload-alt"></i>
                    <p>点击或拖拽Excel文件到此处上传</p>
                    <p style="font-size: 0.85em; color: #999;">支持 .xlsx, .xls 格式</p>
                    <input type="file" id="excel_files" name="excel_files" multiple accept=".xlsx,.xls">
                </div>
                <div id="file-list" class="file-list"></div>
            </div>
            <div style="text-align: center;">
                <button type="submit" class="btn btn-success" id="submit-btn">
                    <i class="fas fa-play"></i> 开始处理
                </button>
                <button type="button" class="btn btn-danger" onclick="clearFiles()">
                    <i class="fas fa-trash"></i> 清空文件
                </button>
            </div>
        </form>
        <div id="progress" class="progress">
            <div class="spinner"></div>
            <p style="text-align: center; color: #667eea;">正在处理，请稍候...</p>
        </div>
        <div id="result" class="result"></div>
    </div>
    <script>
        let selectedFiles = [];
        document.addEventListener('DOMContentLoaded', function() {
            const dropZone = document.getElementById('drop-zone');
            const fileInput = document.getElementById('excel_files');
            const fileList = document.getElementById('file-list');
            fileList.innerHTML = '<div class="empty-state">暂无文件，请选择Excel文件</div>';
            dropZone.addEventListener('dragover', function(e) {
                e.preventDefault();
                dropZone.classList.add('dragover');
            });
            dropZone.addEventListener('dragleave', function(e) {
                e.preventDefault();
                dropZone.classList.remove('dragover');
            });
            dropZone.addEventListener('drop', function(e) {
                e.preventDefault();
                dropZone.classList.remove('dragover');
                handleFiles(e.dataTransfer.files);
            });
            fileInput.addEventListener('change', function(e) {
                handleFiles(e.target.files);
            });
            document.getElementById('upload-form').addEventListener('submit', function(e) {
                e.preventDefault();
                if (selectedFiles.length === 0) {
                    alert('请选择至少一个Excel文件');
                    return;
                }
                document.getElementById('progress').style.display = 'block';
                document.getElementById('result').style.display = 'none';
                document.getElementById('submit-btn').disabled = true;
                const formData = new FormData();
                selectedFiles.forEach(file => {
                    formData.append('excel_files', file);
                });
                fetch('/process', {
                    method: 'POST',
                    body: formData
                })
                .then(response => response.json())
                .then(data => {
                    document.getElementById('progress').style.display = 'none';
                    document.getElementById('submit-btn').disabled = false;
                    if (data.status === 'success') {
                        displayResults(data.results, data.session_id);
                    } else {
                        alert('处理失败: ' + data.message);
                    }
                })
                .catch(error => {
                    document.getElementById('progress').style.display = 'none';
                    document.getElementById('submit-btn').disabled = false;
                    alert('处理失败: ' + error.message);
                });
            });
        });
        function handleFiles(files) {
            for (let i = 0; i < files.length; i++) {
                const file = files[i];
                const ext = file.name.toLowerCase().substring(file.name.lastIndexOf('.'));
                if (!['.xlsx', '.xls'].includes(ext)) {
                    alert('文件 ' + file.name + ' 不是Excel格式');
                    continue;
                }
                const isDuplicate = selectedFiles.some(f => f.name === file.name && f.size === file.size);
                if (!isDuplicate) {
                    selectedFiles.push(file);
                }
            }
            updateFileList();
        }
        function updateFileList() {
            const fileList = document.getElementById('file-list');
            if (selectedFiles.length === 0) {
                fileList.innerHTML = '<div class="empty-state">暂无文件，请选择Excel文件</div>';
                return;
            }
            let html = '';
            selectedFiles.forEach((file, index) => {
                const sizeMB = (file.size / (1024 * 1024)).toFixed(2);
                html += '<div class="file-item"><div class="file-info"><i class="fas fa-file-excel"></i><span class="file-name">' + file.name + '</span><span class="file-size">' + sizeMB + ' MB</span></div><button class="remove-file" onclick="removeFile(' + index + ')"><i class="fas fa-times"></i></button></div>';
            });
            fileList.innerHTML = html;
        }
        function removeFile(index) {
            selectedFiles.splice(index, 1);
            updateFileList();
        }
        function clearFiles() {
            selectedFiles = [];
            document.getElementById('excel_files').value = '';
            updateFileList();
            document.getElementById('result').style.display = 'none';
        }
        function displayResults(results, sessionId) {
            const resultDiv = document.getElementById('result');
            let successCount = 0;
            let noImageCount = 0;
            let failCount = 0;
            let html = '';
            results.forEach(result => {
                if (result.status === 'success') {
                    successCount++;
                    html += '<div class="result-card"><div class="result-title"><i class="fas fa-check-circle"></i> ' + result.filename + '</div>';
                    if (result.tracking_number) {
                        html += '<div class="result-item"><i class="fas fa-barcode"></i> Tracking Number: ' + result.tracking_number + '</div>';
                    }
                    html += '<div class="result-item"><i class="fas fa-image"></i> 提取图片数量: ' + result.image_count + '</div><div class="result-item"><i class="fas fa-file-pdf"></i> 生成PDF: ' + result.pdf_filename + '</div></div>';
                } else if (result.status === 'no_images') {
                    noImageCount++;
                    html += '<div class="result-card warning"><div class="result-title"><i class="fas fa-exclamation-triangle"></i> ' + result.filename + '</div><div class="result-item">' + result.message + '</div></div>';
                } else {
                    failCount++;
                    html += '<div class="result-card error"><div class="result-title"><i class="fas fa-times-circle"></i> ' + result.filename + '</div><div class="result-item">' + result.message + '</div></div>';
                }
            });
            if (successCount > 0) {
                html += '<div class="download-section"><a href="/download/' + sessionId + '" class="btn btn-success"><i class="fas fa-download"></i> 下载所有PDF文件</a></div>';
            }
            resultDiv.innerHTML = html;
            resultDiv.style.display = 'block';
        }
    </script>
</body>
</html>
'''

app = Flask(__name__)
app.secret_key = 'excel-to-pdf-secret-key'
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/process', methods=['POST'])
def process():
    try:
        if 'excel_files' not in request.files:
            return jsonify({'status': 'error', 'message': '没有上传文件'})
        files = request.files.getlist('excel_files')
        if not files or all(f.filename == '' for f in files):
            return jsonify({'status': 'error', 'message': '没有选择文件'})
        session_id = str(uuid.uuid4())
        upload_dir = os.path.join(UPLOAD_FOLDER, session_id)
        os.makedirs(upload_dir, exist_ok=True)
        excel_paths = []
        for file in files:
            if file.filename != '':
                filename = file.filename
                file_path = os.path.join(upload_dir, filename)
                file.save(file_path)
                excel_paths.append(file_path)
        results, session_id = process_excel_files(excel_paths)
        shutil.rmtree(upload_dir)
        return jsonify({
            'status': 'success',
            'results': results,
            'session_id': session_id
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/download/<session_id>')
def download(session_id):
    try:
        processed_dir = os.path.join(PROCESSED_FOLDER, session_id)
        if not os.path.exists(processed_dir):
            return "文件不存在", 404
        pdf_files = [f for f in os.listdir(processed_dir) if f.endswith('.pdf')]
        if len(pdf_files) == 0:
            return "没有PDF文件", 404
        elif len(pdf_files) == 1:
            pdf_path = os.path.join(processed_dir, pdf_files[0])
            return send_file(pdf_path, as_attachment=True, download_name=pdf_files[0])
        else:
            zip_path = os.path.join(PROCESSED_FOLDER, f"pdf_files_{session_id}.zip")
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for pdf_file in pdf_files:
                    pdf_path = os.path.join(processed_dir, pdf_file)
                    zipf.write(pdf_path, pdf_file)
            return send_file(zip_path, as_attachment=True, download_name=f"excel_images_{session_id}.zip")
    except Exception as e:
        return f"下载失败: {str(e)}", 500

class DesktopApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel图片提取转PDF工具")
        self.root.geometry("500x380")
        self.root.resizable(False, False)
        
        self.port = find_available_port()
        self.local_ip = get_local_ip()
        self.server_thread = None
        self.is_running = False
        self.error_msg = None
        
        self.setup_ui()
        self.center_window()
        
    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
    def setup_ui(self):
        style = ttk.Style()
        style.configure('Title.TLabel', font=('Microsoft YaHei UI', 16, 'bold'))
        style.configure('Info.TLabel', font=('Microsoft YaHei UI', 10))
        style.configure('Status.TLabel', font=('Microsoft YaHei UI', 11, 'bold'))
        
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        title_label = ttk.Label(main_frame, text="📊 Excel图片提取转PDF工具", style='Title.TLabel')
        title_label.pack(pady=(0, 20))
        
        info_frame = ttk.LabelFrame(main_frame, text="访问地址", padding="10")
        info_frame.pack(fill=tk.X, pady=(0, 15))
        
        local_url = f"http://127.0.0.1:{self.port}"
        network_url = f"http://{self.local_ip}:{self.port}"
        
        ttk.Label(info_frame, text=f"📍 本地访问: {local_url}", style='Info.TLabel').pack(anchor=tk.W, pady=2)
        ttk.Label(info_frame, text=f"🌐 局域网访问: {network_url}", style='Info.TLabel').pack(anchor=tk.W, pady=2)
        
        self.status_frame = ttk.LabelFrame(main_frame, text="服务状态", padding="10")
        self.status_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.status_label = ttk.Label(self.status_frame, text="⏹ 未启动", style='Status.TLabel')
        self.status_label.pack()
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.start_btn = ttk.Button(btn_frame, text="🚀 启动服务", command=self.start_server, width=15)
        self.start_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.stop_btn = ttk.Button(btn_frame, text="⏹ 停止服务", command=self.stop_server, width=15, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.open_btn = ttk.Button(btn_frame, text="🌐 打开浏览器", command=self.open_browser, width=15, state=tk.DISABLED)
        self.open_btn.pack(side=tk.LEFT)
        
        usage_frame = ttk.LabelFrame(main_frame, text="使用说明", padding="10")
        usage_frame.pack(fill=tk.BOTH, expand=True)
        
        usage_text = "1. 点击\"启动服务\"按钮启动Web服务\n2. 点击\"打开浏览器\"或手动访问上述地址\n3. 拖拽或选择Excel文件上传\n4. 系统自动提取图片并生成PDF\n5. 使用完毕后点击\"停止服务\""
        
        usage_label = ttk.Label(usage_frame, text=usage_text, style='Info.TLabel', justify=tk.LEFT)
        usage_label.pack(anchor=tk.W)
        
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
    def run_flask_app(self):
        try:
            from werkzeug.serving import make_server
            self.server = make_server('0.0.0.0', self.port, app, threaded=True)
            self.server.serve_forever()
        except Exception as e:
            self.error_msg = f"服务错误:\n{str(e)}\n\n详细信息:\n{traceback.format_exc()}"
            self.is_running = False
            
    def start_server(self):
        if self.is_running:
            return
            
        self.is_running = True
        self.error_msg = None
        self.status_label.config(text="⏳ 正在启动...")
        self.start_btn.config(state=tk.DISABLED)
        self.root.update()
        
        try:
            self.server_thread = threading.Thread(target=self.run_flask_app, daemon=True)
            self.server_thread.start()
            
            for i in range(10):
                time.sleep(0.5)
                if self.error_msg:
                    raise Exception(self.error_msg)
                if self.server_thread.is_alive() and hasattr(self, 'server'):
                    break
            
            if self.server_thread.is_alive():
                self.status_label.config(text="✅ 运行中")
                self.stop_btn.config(state=tk.NORMAL)
                self.open_btn.config(state=tk.NORMAL)
                
                self.root.after(1000, self.open_browser)
            else:
                raise Exception(self.error_msg or "服务启动失败，请检查端口是否被占用")
                
        except Exception as e:
            error_text = str(e)
            if len(error_text) > 500:
                error_text = error_text[:500] + "..."
            messagebox.showerror("启动失败", error_text)
            self.is_running = False
            self.status_label.config(text="⏹ 未启动")
            self.start_btn.config(state=tk.NORMAL)
            
    def stop_server(self):
        if self.is_running:
            try:
                if hasattr(self, 'server'):
                    self.server.shutdown()
            except:
                pass
                    
            self.server_thread = None
            self.is_running = False
            self.status_label.config(text="⏹ 已停止")
            self.start_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            self.open_btn.config(state=tk.DISABLED)
            
    def open_browser(self):
        url = f"http://127.0.0.1:{self.port}"
        webbrowser.open(url)
        
    def on_closing(self):
        if self.is_running:
            if messagebox.askokcancel("退出", "服务正在运行，确定要退出吗？"):
                self.stop_server()
                self.root.destroy()
        else:
            self.root.destroy()
            
    def run(self):
        self.root.mainloop()

if __name__ == '__main__':
    desktop_app = DesktopApp()
    desktop_app.run()
