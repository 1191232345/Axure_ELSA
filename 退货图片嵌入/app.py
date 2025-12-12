from flask import Flask, request, jsonify, send_file, render_template_string
import os
import zipfile
import tempfile
import shutil
import re
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
import uuid
import socket
import webbrowser
import threading
from werkzeug.utils import secure_filename

app = Flask(__name__)

# 获取当前工作目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
PROCESSED_FOLDER = os.path.join(BASE_DIR, 'processed')

print(f"项目根目录: {BASE_DIR}")
print(f"上传目录: {UPLOAD_FOLDER}")
print(f"处理目录: {PROCESSED_FOLDER}")

# 确保目录存在的函数
def ensure_directories():
    """确保所有必要的目录都存在"""
    try:
        if not os.path.exists(UPLOAD_FOLDER):
            os.makedirs(UPLOAD_FOLDER)
            print(f"✅ 创建上传目录: {UPLOAD_FOLDER}")
        else:
            print(f"✅ 上传目录已存在: {UPLOAD_FOLDER}")
        
        if not os.path.exists(PROCESSED_FOLDER):
            os.makedirs(PROCESSED_FOLDER)
            print(f"✅ 创建处理目录: {PROCESSED_FOLDER}")
        else:
            print(f"✅ 处理目录已存在: {PROCESSED_FOLDER}")
        
        return True
    except Exception as e:
        print(f"❌ 创建目录失败: {e}")
        return False

# 启动时确保目录存在
ensure_directories()

def extract_elsa_from_zipname(zip_filename):
    """从压缩包文件名提取ELSA跟踪号（取首个-之后的数据）"""
    name_without_ext = os.path.splitext(zip_filename)[0]
    
    # 取首个-之后的数据，如果没有-则使用整个文件名
    if '-' in name_without_ext:
        main_part = name_without_ext.split('-', 1)[1].strip()  # 取第一个-之后的部分
    else:
        main_part = name_without_ext
    
    # 只保留字母和数字
    clean_elsa = re.sub(r'[^a-zA-Z0-9]', '', main_part)
    return clean_elsa

def clean_elsa_tracking(elsa_value):
    """清理ELSA跟踪号，只保留字母和数字"""
    if not elsa_value:
        return ""
    return re.sub(r'[^a-zA-Z0-9]', '', str(elsa_value))

def parse_excel_data(excel_path):
    """解析Excel文件，提取客户、ELSA跟踪号、RFID信息"""
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 查找各列的索引
    header_row = 1
    customer_col = None
    elsa_col = None
    rfid_col = None
    
    # 可能的列名
    customer_keywords = ['客户', 'customer', '客户代码', '客户编号']
    elsa_keywords = ['ELSA', 'elsa', '跟踪号', '追踪号', 'tracking']
    rfid_keywords = ['RFID', 'rfid', '射频', '标签']
    
    for col_idx in range(1, ws.max_column + 1):
        cell_value = str(ws.cell(row=header_row, column=col_idx).value or '')
        cell_lower = cell_value.lower()
        
        if any(keyword in cell_lower for keyword in customer_keywords):
            customer_col = col_idx
        elif any(keyword in cell_lower for keyword in elsa_keywords):
            elsa_col = col_idx
        elif any(keyword in cell_lower for keyword in rfid_keywords):
            rfid_col = col_idx
    
    # 设置默认列
    if customer_col is None: customer_col = 1
    if elsa_col is None: elsa_col = 2
    if rfid_col is None: rfid_col = 3
    
    excel_data = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        customer = str(ws.cell(row=row_idx, column=customer_col).value or '')
        elsa_tracking = str(ws.cell(row=row_idx, column=elsa_col).value or '')
        rfid = str(ws.cell(row=row_idx, column=rfid_col).value or '')
        
        if any([customer.strip(), elsa_tracking.strip(), rfid.strip()]):
            excel_data.append({
                'row_idx': row_idx,
                'customer': customer.strip(),
                'elsa_tracking': elsa_tracking.strip(),
                'clean_elsa': clean_elsa_tracking(elsa_tracking),
                'rfid': rfid.strip()
            })
    
    return excel_data, customer_col, elsa_col, rfid_col

# HTML模板
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>ELSA跟踪号匹配工具</title>
    <meta charset="utf-8">
    <style>
        body { 
            font-family: "Microsoft YaHei", Arial, sans-serif; 
            max-width: 1200px; 
            margin: 50px auto; 
            padding: 20px; 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }
        .container {
            background: white;
            padding: 40px;
            border-radius: 15px;
            box-shadow: 0 15px 35px rgba(0,0,0,0.1);
        }
        h1 {
            color: #333;
            text-align: center;
            margin-bottom: 30px;
            font-weight: 300;
            font-size: 2.5em;
        }
        .form-group { 
            margin: 20px 0; 
        }
        label {
            display: block;
            margin-bottom: 8px;
            font-weight: bold;
            color: #555;
        }
        input[type="file"] {
            width: 100%;
            padding: 15px;
            border: 2px dashed #ccc;
            border-radius: 8px;
            background: #f9f9f9;
            cursor: pointer;
            transition: all 0.3s ease;
        }
        input[type="file"]:hover {
            border-color: #667eea;
            background: #f0f4ff;
        }
        button {
            padding: 15px 30px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 16px;
            margin: 10px 0;
            transition: all 0.3s ease;
        }
        button:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(102, 126, 234, 0.4);
        }
        button.secondary {
            background: linear-gradient(135deg, #6c757d 0%, #495057 100%);
        }
        button.warning {
            background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        }

        /* 弹窗样式 */
        .modal {
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.5);
        }
        .modal-content {
            background-color: white;
            margin: 2% auto;
            padding: 20px;
            border-radius: 10px;
            width: 95%;
            max-width: 1400px;
            max-height: 90vh;
            overflow-y: auto;
        }
        .close {
            color: #aaa;
            float: right;
            font-size: 28px;
            font-weight: bold;
            cursor: pointer;
        }
        .close:hover {
            color: black;
        }
        .match-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            font-size: 14px;
            table-layout: auto; /* 允许列宽根据内容调整 */
        }
        .match-table th, .match-table td {
            border: 1px solid #ddd;
            padding: 8px 12px;
            text-align: left;
            white-space: nowrap;
        }
        .match-table th {
            background-color: #f8f9fa;
            font-weight: bold;
            position: sticky;
            top: 0;
        }
        .match-table tr:hover {
            background-color: #f5f5f5;
        }
        .input-field {
            width: 100%;
            padding: 6px;
            border: 1px solid #ddd;
            border-radius: 4px;
            font-size: 14px;
            box-sizing: border-box;
        }
        .search-box {
            margin-bottom: 15px;
            padding: 10px;
            width: 100%;
            border: 1px solid #ddd;
            border-radius: 4px;
            font-size: 16px;
        }
        .table-container {
            max-height: 60vh;
            overflow-y: auto;
        }
        .match-status {
            padding: 4px 8px;
            border-radius: 4px;
            font-size: 12px;
        }
        .matched {
            background-color: #d4edda;
            color: #155724;
        }
        .unmatched {
            background-color: #f8d7da;
            color: #721c24;
        }
        
        /* 文件列表样式优化 */
        .file-list {
            border: 1px solid #ddd;
            border-radius: 4px;
            max-height: 150px;
            overflow-y: auto;
            padding: 10px;
            background-color: #f9f9f9;
        }
        .file-item {
            padding: 5px 0;
            border-bottom: 1px solid #eee;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .file-item:last-child {
            border-bottom: none;
        }
        .remove-btn {
            background: #dc3545;
            color: white;
            border: none;
            padding: 2px 8px;
            border-radius: 3px;
            cursor: pointer;
            font-size: 12px;
        }
        .remove-btn:hover {
            background: #c82333;
        }
        
        /* 下拉框样式 */
        select.input-field {
            appearance: auto;
            background: white;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>📋 ELSA跟踪号匹配工具</h1>
        
        <div class="info">
            <strong>使用说明：</strong>
            <ol>
                <li>上传一个或多个ZIP压缩包</li>
                <li>上传Excel文件（需要包含客户、ELSA跟踪号、RFID列）</li>
                <li>系统会自动提取压缩包名称中首个-之后的ELSA跟踪号</li>
                <li>与Excel中的ELSA跟踪号进行完全匹配</li>
                <li>在弹窗中确认或修改匹配结果</li>
                <li>导出匹配结果表格</li>
            </ol>
        </div>

        <form id="uploadForm" enctype="multipart/form-data">
            <div class="form-group">
                <label>📦 ZIP压缩包文件 (可多选):</label>
                <input type="file" name="zip_files" accept=".zip" multiple required>
                <div class="file-list" id="zipFileList">
                    <div>暂无文件，请选择ZIP文件</div>
                </div>
            </div>
            <div class="form-group">
                <label>📊 Excel文件:</label>
                <input type="file" name="excel_file" accept=".xlsx,.xls" required>
                <div class="file-list" id="excelFileList">
                    <div>暂无文件，请选择Excel文件</div>
                </div>
            </div>
            
            <button type="button" onclick="processFiles()">🚀 开始匹配</button>
            <button type="button" onclick="clearFiles()" class="secondary">🗑️ 清空文件</button>
        </form>

        <div class="loading" id="loading">
            <div class="spinner"></div>
            <p>处理中，请稍候...</p>
        </div>

        <div class="result" id="result">
            <!-- 结果内容将由JavaScript动态填充 -->
        </div>
    </div>

    <!-- 匹配结果弹窗 -->
    <div id="matchModal" class="modal">
        <div class="modal-content">
            <span class="close" onclick="closeModal()">&times;</span>
            <h2>📋 ELSA跟踪号匹配结果</h2>
            <input type="text" id="searchBox" class="search-box" placeholder="🔍 搜索客户、ELSA跟踪号或RFID..." onkeyup="filterTable()">
            
            <div class="table-container">
                <table class="match-table" id="matchTable">
                    <thead>
                        <tr>
                            <th>行号</th>
                            <th>客户</th>
                            <th>ELSA跟踪号</th>
                            <th>RFID</th>
                            <th>匹配状态</th>
                            <th>压缩包名称</th>
                        </tr>
                    </thead>
                    <tbody id="matchTableBody">
                        <!-- 表格内容将由JavaScript动态填充 -->
                    </tbody>
                </table>
            </div>
            
            <div style="margin-top: 20px; text-align: center;">
                <button onclick="exportMatches()">💾 导出匹配表</button>
                <button onclick="exportWithImages()" class="warning">🖼️ 导出带图片匹配表</button>
                <button onclick="closeModal()" class="secondary">❌ 关闭</button>
            </div>
        </div>
    </div>

    <script>
        let selectedZipFiles = [];
        let selectedExcelFile = null;
        let matchData = [];
        let zipFilesList = [];

        // 初始化隐藏加载和结果区域
        document.getElementById('loading').style.display = 'none';
        document.getElementById('result').style.display = 'none';
        document.getElementById('matchModal').style.display = 'none';

        // 初始化文件选择监听
        document.querySelector('input[name="zip_files"]').addEventListener('change', function(e) {
            updateFileList(e.target.files, 'zip');
        });

        document.querySelector('input[name="excel_file"]').addEventListener('change', function(e) {
            updateFileList(e.target.files, 'excel');
        });

        function updateFileList(files, type) {
            const fileList = document.getElementById(type + 'FileList');
            
            if (!files || files.length === 0) {
                fileList.innerHTML = '<div>暂无文件，请选择' + (type === 'zip' ? 'ZIP' : 'Excel') + '文件</div>';
                if (type === 'zip') {
                    selectedZipFiles = [];
                } else {
                    selectedExcelFile = null;
                }
                return;
            }

            let html = '';
            if (type === 'zip') {
                selectedZipFiles = Array.from(files);
                selectedZipFiles.forEach((file, index) => {
                    const sizeMB = (file.size / (1024 * 1024)).toFixed(2);
                    html += `
                        <div class="file-item">
                            📦 ${file.name} (${sizeMB} MB)
                            <button class="remove-btn" onclick="removeFile(${index}, 'zip')">移除</button>
                        </div>
                    `;
                });
            } else {
                selectedExcelFile = files[0];
                const sizeMB = (selectedExcelFile.size / (1024 * 1024)).toFixed(2);
                html = `
                    <div class="file-item">
                        📊 ${selectedExcelFile.name} (${sizeMB} MB)
                        <button class="remove-btn" onclick="removeFile(0, 'excel')">移除</button>
                    </div>
                `;
            }
            
            fileList.innerHTML = html;
        }

        function removeFile(index, type) {
            if (type === 'zip') {
                selectedZipFiles.splice(index, 1);
                updateFileList(selectedZipFiles, 'zip');
            } else {
                selectedExcelFile = null;
                document.querySelector('input[name="excel_file"]').value = '';
                updateFileList([], 'excel');
            }
        }

        function clearFiles() {
            selectedZipFiles = [];
            selectedExcelFile = null;
            document.querySelector('input[name="zip_files"]').value = '';
            document.querySelector('input[name="excel_file"]').value = '';
            document.getElementById('zipFileList').innerHTML = '<div>暂无文件，请选择ZIP文件</div>';
            document.getElementById('excelFileList').innerHTML = '<div>暂无文件，请选择Excel文件</div>';
            document.getElementById('result').style.display = 'none';
        }

        function processFiles() {
            if (selectedZipFiles.length === 0 || !selectedExcelFile) {
                showResult('❌ 请选择至少一个ZIP文件和一个Excel文件', 'error');
                return;
            }

            const formData = new FormData();
            
            // 添加所有ZIP文件
            selectedZipFiles.forEach((file, index) => {
                formData.append('zip_files', file);
            });
            
            // 添加Excel文件
            formData.append('excel_file', selectedExcelFile);

            // 显示加载动画
            document.getElementById('loading').style.display = 'block';
            document.getElementById('result').style.display = 'none';

            fetch('/process', {
                method: 'POST',
                body: formData
            })
            .then(response => response.json())
            .then(data => {
                if (data.status === 'success') {
                    matchData = data.match_data || [];
                    zipFilesList = data.zip_files || [];
                    showMatchModal();
                } else {
                    showResult(`
                        <h3>❌ 处理失败</h3>
                        <p>${data.message}</p>
                        <button onclick="resetForm()">🔄 重试</button>
                    `, 'error');
                }
            })
            .catch(error => {
                showResult(`
                    <h3>❌ 网络错误</h3>
                    <p>${error.message}</p>
                    <button onclick="resetForm()">🔄 重试</button>
                `, 'error');
            })
            .finally(() => {
                document.getElementById('loading').style.display = 'none';
            });
        }

        function showMatchModal() {
            const modal = document.getElementById('matchModal');
            const tbody = document.getElementById('matchTableBody');
            
            // 清空表格
            tbody.innerHTML = '';
            
            // 填充表格数据
            matchData.forEach((row, index) => {
                const tr = document.createElement('tr');
                const isMatched = row.is_matched;
                
                // 创建ZIP文件下拉选项
                let zipOptions = '<option value="">未匹配</option>';
                zipFilesList.forEach(zipFile => {
                    const selected = zipFile === row.matched_zip ? 'selected' : '';
                    zipOptions += `<option value="${zipFile}" ${selected}>${zipFile}</option>`;
                });
                
                tr.innerHTML = `
                    <td>${row.row_idx}</td>
                    <td><input type="text" class="input-field" value="${row.customer || ''}" onchange="updateMatchData(${index}, 'customer', this.value)" readonly></td>
                    <td><input type="text" class="input-field" value="${row.excel_elsa || ''}" onchange="updateMatchData(${index}, 'excel_elsa', this.value)" readonly></td>
                    <td><input type="text" class="input-field" value="${row.rfid || ''}" onchange="updateMatchData(${index}, 'rfid', this.value)" readonly></td>
                    <td>
                        <span class="match-status ${isMatched ? 'matched' : 'unmatched'}">
                            ${isMatched ? '✅ 已匹配' : '❌ 未匹配'}
                        </span>
                    </td>
                    <td>
                        <select class="input-field" onchange="updateZipFile(${index}, this.value)">
                            ${zipOptions}
                        </select>
                    </td>
                `;
                
                tbody.appendChild(tr);
            });
            
            modal.style.display = 'block';
        }

        function updateMatchData(index, field, value) {
            if (matchData[index]) {
                matchData[index][field] = value;
            }
        }
        
        function updateZipFile(index, zipFilename) {
            if (matchData[index]) {
                matchData[index].matched_zip = zipFilename;
                matchData[index].is_matched = !!zipFilename;
                
                // 更新匹配状态显示
                const statusCell = document.querySelectorAll('.match-status')[index];
                if (statusCell) {
                    if (zipFilename) {
                        statusCell.className = 'match-status matched';
                        statusCell.textContent = '✅ 已匹配';
                    } else {
                        statusCell.className = 'match-status unmatched';
                        statusCell.textContent = '❌ 未匹配';
                    }
                }
            }
        }

        function closeModal() {
            document.getElementById('matchModal').style.display = 'none';
        }

        function filterTable() {
            const searchText = document.getElementById('searchBox').value.toLowerCase();
            const rows = document.getElementById('matchTableBody').getElementsByTagName('tr');
            
            for (let row of rows) {
                const cells = row.getElementsByTagName('td');
                let matchFound = false;
                
                for (let cell of cells) {
                    if (cell.textContent.toLowerCase().includes(searchText)) {
                        matchFound = true;
                        break;
                    }
                }
                
                row.style.display = matchFound ? '' : 'none';
            }
        }

        function exportMatches() {
            // 发送匹配结果到后端生成Excel
            fetch('/export_matches', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json',
                },
                body: JSON.stringify({
                    matches: matchData
                })
            })
            .then(response => response.json())
            .then(data => {
                if (data.status === 'success') {
                    // 下载文件
                    window.location.href = '/download_export/' + data.filename;
                    closeModal();
                    showResult(`
                        <h3>✅ 导出成功！</h3>
                        <p>已生成匹配结果表格</p>
                        <p><a href="/download_export/${data.filename}" style="color: #2196F3; text-decoration: none; font-weight: bold;">
                           📥 点击下载匹配表
                        </a></p>
                        <button onclick="resetForm()">🔄 继续处理</button>
                    `, 'success');
                } else {
                    alert('导出失败: ' + data.message);
                }
            })
            .catch(error => {
                alert('网络错误: ' + error.message);
            });
        }

        function exportWithImages() {
            // 发送匹配结果和ZIP文件信息到后端生成带图片的Excel
            fetch('/export_with_images', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json',
                },
                body: JSON.stringify({
                    matches: matchData,
                    zip_files: zipFilesList
                })
            })
            .then(response => response.json())
            .then(data => {
                if (data.status === 'success') {
                    // 下载文件
                    window.location.href = '/download_export/' + data.filename;
                    closeModal();
                    showResult(`
                        <h3>✅ 导出成功！</h3>
                        <p>已生成带图片的匹配结果表格</p>
                        <p><a href="/download_export/${data.filename}" style="color: #2196F3; text-decoration: none; font-weight: bold;">
                           📥 点击下载带图片匹配表
                        </a></p>
                        <button onclick="resetForm()">🔄 继续处理</button>
                    `, 'success');
                } else {
                    alert('导出失败: ' + data.message);
                }
            })
            .catch(error => {
                alert('网络错误: ' + error.message);
            });
        }

        function showResult(content, type) {
            const resultDiv = document.getElementById('result');
            resultDiv.innerHTML = content;
            resultDiv.className = 'result ' + (type === 'error' ? 'error' : '');
            resultDiv.style.display = 'block';
        }

        function resetForm() {
            clearFiles();
        }

        // 点击模态框外部关闭
        window.onclick = function(event) {
            const modal = document.getElementById('matchModal');
            if (event.target === modal) {
                closeModal();
            }
        }
    </script>
</body>
</html>
'''

def extract_info_from_zipname(zip_filename):
    """从压缩包文件名提取ELSA跟踪号信息"""
    name_without_ext = os.path.splitext(zip_filename)[0]
    
    # 取最后一个"-"之后的部分，如果没有"-"则使用全名
    if '-' in name_without_ext:
        main_part = name_without_ext.split('-')[-1].strip()
    else:
        main_part = name_without_ext
    
    # 清理特殊字符，只保留字母数字和下划线
    clean_name = re.sub(r'[^a-zA-Z0-9_]', '', main_part)
    
    # 尝试匹配各种ELSA跟踪号格式
    patterns = [
        r'(\d{6,})',  # 6位以上数字
        r'([A-Za-z]{4,})',  # 4位以上英文
        r'(\d+_[A-Za-z]+)',  # 数字_英文
        r'([A-Za-z]+_\d+)',  # 英文_数字
        r'([A-Za-z]+\d+)',  # 英文+数字
        r'(\d+[A-Za-z]+)',  # 数字+英文
        r'ELS[Aa]?([A-Za-z\d_]+)',  # ELSA开头
    ]
    
    for pattern in patterns:
        match = re.search(pattern, main_part)
        if match:
            if 'ELS[Aa]?' in pattern:
                return match.group(1)
            else:
                return match.group(0)
    
    return clean_name

def find_elsa_match(zip_elsa, excel_data):
    """基于ELSA跟踪号进行匹配"""
    if not zip_elsa:
        return None, "无ELSA跟踪号"
    
    for excel_row in excel_data:
        excel_elsa = excel_row['elsa_tracking']
        clean_excel_elsa = excel_row['clean_elsa']
        
        if not excel_elsa:
            continue
        
        if zip_elsa == clean_excel_elsa:
            return excel_row, "完全匹配"
        
        if zip_elsa in clean_excel_elsa or clean_excel_elsa in zip_elsa:
            return excel_row, "部分匹配"
    
    return None, "未找到匹配"

def process_images_from_zips(zip_paths, excel_path, output_excel_path, manual_matches=None):
    """处理ZIP文件中的图片并插入到Excel，保持原表格格式"""
    zip_info_list = []
    temp_dirs = []
    match_stats = {
        'total_zips': len(zip_paths),
        'matched_zips': 0,
        'total_images': 0,
        'match_details': [],
        'match_details_list': []
    }
    
    try:
        ensure_directories()
        # 正确解包所有返回值
        excel_data, customer_col, elsa_col, rfid_col = parse_excel_data(excel_path)
        
        # 解析压缩包信息
        for zip_path in zip_paths:
            zip_filename = os.path.basename(zip_path)
            elsa_tracking = extract_elsa_from_zipname(zip_filename)
            zip_info = {
                'path': zip_path,
                'filename': zip_filename,
                'elsa_tracking': elsa_tracking,
                'images': []
            }
            zip_info_list.append(zip_info)
        
        # 解压并收集图片
        for zip_info in zip_info_list:
            temp_dir = tempfile.mkdtemp()
            temp_dirs.append(temp_dir)
            
            try:
                with zipfile.ZipFile(zip_info['path'], 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)
                
                for root, dirs, files in os.walk(temp_dir):
                    for file in files:
                        if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                            zip_info['images'].append(os.path.join(root, file))
                
                match_stats['total_images'] += len(zip_info['images'])
                
            except Exception as e:
                continue
        
        # 处理Excel - 保持原有格式
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 查找图片插入起始列（在最后一列之后）
        image_col_idx = ws.max_column + 1
        
        processed_count = 0
        
        # 创建ZIP文件名到信息的映射，便于查找
        zip_name_mapping = {zip_info['filename']: zip_info for zip_info in zip_info_list}
        
        for match in manual_matches:
            zip_filename = match['zip_filename']
            if zip_filename not in zip_name_mapping:
                continue
                
            zip_info = zip_name_mapping[zip_filename]
            row_idx = match['excel_row']
            
            # 插入图片，每张图片独占一个单元格
            for i, img_path in enumerate(zip_info['images']):
                try:
                    # 计算图片应该插入的单元格位置
                    target_col = image_col_idx + i
                    target_row = row_idx
                    
                    # 确保列标题存在
                    if ws.cell(row=1, column=target_col).value is None:
                        ws.cell(row=1, column=target_col, value=f'图片{i+1}')
                    
                    # 调整列宽
                    ws.column_dimensions[get_column_letter(target_col)].width = 15
                    
                    # 插入图片
                    img = Image(img_path)
                    img.width = 80
                    img.height = 80
                    
                    # 将图片锚定到特定单元格（确保不重叠）
                    cell_address = f"{get_column_letter(target_col)}{target_row}"
                    img.anchor = cell_address
                    ws.add_image(img)
                    
                    processed_count += 1
                    match_stats['matched_zips'] += 1
                except Exception as e:
                    print(f"插入图片错误: {e}")
                    continue
        
        # 调整行高以适应图片
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 70  # 调整行高以适应图片
        
        ensure_directories()
        wb.save(output_excel_path)
        
        return True, processed_count, len(zip_paths), 'ELSA跟踪号匹配', match_stats, match_stats['matched_zips']
        
    except Exception as e:
        print(f"处理错误: {e}")
        import traceback
        traceback.print_exc()
        return False, 0, 0, '错误', match_stats, 0
    finally:
        for temp_dir in temp_dirs:
            try:
                shutil.rmtree(temp_dir, ignore_errors=True)
            except:
                pass

def get_local_ip():
    """获取本地IP地址"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "127.0.0.1"

def open_browser():
    """自动打开浏览器"""
    import time
    time.sleep(1.5)
    webbrowser.open("http://localhost:5000")

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/process', methods=['POST'])
def process_files():
    try:
        ensure_directories()
        
        if 'zip_files' not in request.files or 'excel_file' not in request.files:
            return jsonify({'status': 'error', 'message': '请上传ZIP文件和Excel文件'})
        
        zip_files = request.files.getlist('zip_files')
        excel_file = request.files['excel_file']
        
        if not any(zip_files) or excel_file.filename == '':
            return jsonify({'status': 'error', 'message': '请选择有效的文件'})
        
        # 保存上传的文件
        zip_paths = []
        zip_filenames = []
        for zip_file in zip_files:
            if zip_file and zip_file.filename.endswith('.zip'):
                filename = secure_filename(zip_file.filename)
                zip_path = os.path.join(UPLOAD_FOLDER, filename)
                zip_file.save(zip_path)
                zip_paths.append(zip_path)
                zip_filenames.append(filename)
        
        if not zip_paths:
            return jsonify({'status': 'error', 'message': '没有有效的ZIP文件'})
        
        excel_filename = secure_filename(excel_file.filename)
        excel_path = os.path.join(UPLOAD_FOLDER, excel_filename)
        excel_file.save(excel_path)
        
        # 解析Excel数据
        excel_data, customer_col, elsa_col, rfid_col = parse_excel_data(excel_path)
        
        # 提取压缩包中的ELSA跟踪号
        zip_elsa_mapping = {}
        for zip_filename in zip_filenames:
            elsa_from_zip = extract_elsa_from_zipname(zip_filename)
            zip_elsa_mapping[elsa_from_zip] = zip_filename
        
        # 进行完全匹配
        match_results = []
        for excel_row in excel_data:
            excel_clean_elsa = excel_row['clean_elsa']
            matched_zip = None
            is_matched = False
            
            # 完全匹配逻辑
            if excel_clean_elsa in zip_elsa_mapping:
                matched_zip = zip_elsa_mapping[excel_clean_elsa]
                is_matched = True
            
            match_results.append({
                'row_idx': excel_row['row_idx'],
                'customer': excel_row['customer'],
                'excel_elsa': excel_row['elsa_tracking'],
                'rfid': excel_row['rfid'],
                'zip_elsa': excel_clean_elsa if is_matched else '',
                'matched_zip': matched_zip,
                'is_matched': is_matched
            })
        
        return jsonify({
            'status': 'success',
            'match_data': match_results,
            'zip_files': zip_filenames,
            'message': '匹配完成，请确认结果'
        })
            
    except Exception as e:
        print(f"处理错误: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'服务器错误: {str(e)}'})

@app.route('/export_matches', methods=['POST'])
def export_matches():
    try:
        data = request.get_json()
        matches = data.get('matches', [])
        
        # 创建导出文件
        output_filename = f"match_result_{uuid.uuid4().hex[:8]}.xlsx"
        output_path = os.path.join(PROCESSED_FOLDER, output_filename)
        
        # 创建新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "匹配结果"
        
        # 添加表头
        headers = ["行号", "客户", "Excel ELSA跟踪号", "RFID", "压缩包ELSA跟踪号", "压缩包名称", "匹配状态"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # 添加数据
        for row_idx, match in enumerate(matches, 2):
            ws.cell(row=row_idx, column=1, value=match.get('row_idx', ''))
            ws.cell(row=row_idx, column=2, value=match.get('customer', ''))
            ws.cell(row=row_idx, column=3, value=match.get('excel_elsa', ''))
            ws.cell(row=row_idx, column=4, value=match.get('rfid', ''))
            ws.cell(row=row_idx, column=5, value=match.get('zip_elsa', ''))
            ws.cell(row=row_idx, column=6, value=match.get('matched_zip', ''))
            ws.cell(row=row_idx, column=7, value="已匹配" if match.get('is_matched') else "未匹配")
        
        # 保存文件
        ensure_directories()
        wb.save(output_path)
        
        return jsonify({
            'status': 'success',
            'filename': output_filename,
            'message': '导出成功'
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/export_with_images', methods=['POST'])
def export_with_images():
    """导出包含嵌入图片的表格，保持原表格格式"""
    try:
        data = request.get_json()
        matches = data.get('matches', [])
        zip_files = data.get('zip_files', [])
        
        # 创建导出文件
        output_filename = f"match_result_with_images_{uuid.uuid4().hex[:8]}.xlsx"
        output_path = os.path.join(PROCESSED_FOLDER, output_filename)
        
        # 创建ZIP文件路径列表
        zip_paths = []
        for zip_filename in zip_files:
            zip_path = os.path.join(UPLOAD_FOLDER, zip_filename)
            if os.path.exists(zip_path):
                zip_paths.append(zip_path)
        
        # 使用原始上传的Excel文件作为模板
        # 在process_files中保存了excel文件，这里直接使用
        excel_path = None
        for filename in os.listdir(UPLOAD_FOLDER):
            if filename.endswith(('.xlsx', '.xls')) and not filename.startswith('match_result'):
                excel_path = os.path.join(UPLOAD_FOLDER, filename)
                break
        
        if not excel_path or not os.path.exists(excel_path):
            return jsonify({'status': 'error', 'message': '未找到Excel文件'})
        
        # 准备匹配信息用于图片插入
        manual_matches = []
        for match in matches:
            if match.get('is_matched') and match.get('matched_zip'):
                manual_matches.append({
                    'zip_filename': match.get('matched_zip'),
                    'excel_row': match.get('row_idx'),
                    'excel_elsa': match.get('excel_elsa'),
                    'matched': True
                })
        
        # 处理图片并生成Excel
        success, processed_count, total_zips, message, match_stats, matched_zips = process_images_from_zips(
            zip_paths, 
            excel_path,
            output_path, 
            manual_matches
        )
        
        if success:
            return jsonify({
                'status': 'success',
                'filename': output_filename,
                'message': f'导出成功，处理了{processed_count}张图片'
            })
        else:
            return jsonify({'status': 'error', 'message': '处理失败'})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/download_export/<filename>')
def download_export(filename):
    file_path = os.path.join(PROCESSED_FOLDER, filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True, download_name=f"ELSA匹配结果_{filename}")
    else:
        return jsonify({'status': 'error', 'message': '文件不存在'})

if __name__ == '__main__':
    ensure_directories()
    local_ip = get_local_ip()
    port = 8080
    
    print("=" * 80)
    print("🚀 智能图片处理工具已启动")
    print("=" * 80)
    print(f"📍 本地访问: http://localhost:{port}")
    print(f"🌐 局域网访问: http://{local_ip}:{port}")
    print("=" * 80)
    
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host='0.0.0.0', port=port, debug=True)